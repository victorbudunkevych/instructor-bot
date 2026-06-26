# bot.py - ВЕРСІЯ 20 PRODUCTION
# ВИПРАВЛЕННЯ: rate_student_menu тепер показує всі completed уроки з оцінками - ТЕСТОВА ВЕРСІЯ З ОКРЕМОЮ БД
import sqlite3
import re
import logging
import os
from datetime import datetime, timedelta
from contextlib import contextmanager
from io import BytesIO

from telegram import Update, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardRemove
from telegram.ext import (
    ApplicationBuilder, 
    CommandHandler, 
    MessageHandler, 
    ContextTypes, 
    CallbackQueryHandler,
    filters
)
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ==================== PRODUCTION КОНФІГУРАЦІЯ ====================
# PRODUCTION БОТ TOKEN
TOKEN = "8593442263:AAG6hcvZ_xRdsSoDKade5LMbMdX2MUq4dIA"
ADMIN_ID = [
    669706811,   # Віктор (власник)
    280240917,   # Шепшелей Владислав
    648021272,   # Кузенко Руслана
    884453802    # Стефанюк Ірина
]
TIMEZONE = "Europe/Kyiv"

# БАЗА ДАНИХ НА PERSISTENT DISK
import os
if os.path.exists("/var/data"):
    DB_NAME = "/var/data/driving_school.db"
    print("✅ Використовую Persistent Disk: /var/data/driving_school.db")
else:
    DB_NAME = "driving_school.db"
    print("⚠️ Persistent Disk не знайдено, використовую локальну БД")
# ==================================================================

# Робочі години
WORK_HOURS_START = 8
WORK_HOURS_END = 18
# Ціни за годину
PRICES = {
    "1 година": 420,
    "2 години": 840
}
from database import (
    init_db, 
    init_lessons_table, 
    init_students_table,
    migrate_database,
    get_instructors_by_transmission,
    get_instructor_by_name,
    get_instructor_by_telegram_id,
    get_instructor_rating,
    get_db as _original_get_db,
    init_schedule_blocks_table,
    get_instructor_stats_period,
    get_admin_report_by_instructors,
    get_instructor_report,
    get_all_instructors,
    register_student,
    get_student_by_telegram_id,
    update_lesson,
    add_lesson_rating
)

# ======================= HELPER FUNCTIONS =======================
def get_student_by_phone(phone):
    """Знайти учня за номером телефону"""
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            digits = ''.join(filter(str.isdigit, phone))[-9:]
            cursor.execute("""
                SELECT id, name, phone, tariff, registered_via, telegram_id
                FROM students
                WHERE REPLACE(REPLACE(REPLACE(phone, '+', ''), '-', ''), ' ', '') LIKE ?
            """, (f'%{digits}',))
            return cursor.fetchone()
    except Exception as e:
        logger.error(f"Error in get_student_by_phone: {e}")
        return None

def add_instructor_rating(lesson_id, rating, feedback=""):
    """Додати оцінку та коментар інструктора для учня"""
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE lessons
                SET instructor_rating = ?, instructor_feedback = ?
                WHERE id = ?
            """, (rating, feedback, lesson_id))
            conn.commit()
            return True
    except Exception as e:
        logger.error(f"Error in add_instructor_rating: {e}", exc_info=True)
        return False

# Перевизначаємо get_db для тестової БД
@contextmanager
def get_db():
    """Context manager для роботи з тестовою БД"""
    conn = sqlite3.connect(DB_NAME)
    try:
        yield conn
    finally:
        conn.close()

# Monkey patch для database модуля щоб використовував тестову БД
import database
database.DB_NAME = DB_NAME
database.get_db = get_db

# Налаштування логування
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO,
    handlers=[
        logging.FileHandler('bot.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

TZ = pytz.timezone(TIMEZONE)

# ======================= AUTO-ADD INSTRUCTORS =======================
def ensure_instructors_exist():
    """Автоматично додає інструкторів якщо їх немає в базі"""
    instructors = [
        (5077103081, 'Фірсов Артур', '+380666619757', 'Механіка', 550),
        (197658460, 'Урядко Артур', '+380502380725', 'Автомат', 550),
        (765241025, 'Козюля Ксенія', '+380951750958', 'Автомат', 550),
        (573133979, 'Максим Белей', '+380983203215', 'Автомат', 490),
        (669706811, 'Тест Тест', '+380936879999', 'Автомат', 490),
        (2042857396, 'Будункевич Мирослав', '+380982534001', 'Механіка', 490),
        (7115781216, 'Нагорний Віталій', '+380502994424', 'Механіка', 550),
        (1846725989, 'Рекетчук Богдан', '+380501591448', 'Механіка', 550),
        (7996066111, 'Щербина Василь', '+380950732059', 'Механіка', 550),
        (831664827, 'Данилишин Святослав', '+380960755539', 'Механіка', 550)
    ]
    
    with get_db() as conn:
        cursor = conn.cursor()
        added = 0
        
        for telegram_id, name, phone, transmission, price in instructors:
            cursor.execute("SELECT id FROM instructors WHERE telegram_id = ?", (telegram_id,))
            if not cursor.fetchone():
                cursor.execute("""
                    INSERT INTO instructors (telegram_id, name, phone, transmission_type, price_per_hour, is_active, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (telegram_id, name, phone, transmission, price, 1, datetime.now()))
                logger.info(f"✅ Додано інструктора: {name} ({transmission})")
                added += 1
        
        if added > 0:
            conn.commit()
            logger.info(f"🎉 Автоматично додано {added} інструкторів")
        else:
            logger.info("ℹ️ Всі інструктори вже є в базі")

def is_instructor(telegram_id):
    """Перевіряє чи є користувач інструктором"""
    instructor = get_instructor_by_telegram_id(telegram_id)
    return instructor is not None

# ======================= HELPERS =======================
def get_next_dates(days=14, instructor_name=None):
    """Генерує список дат на найближчі N днів з кількістю вільних годин"""
    dates = []
    now = datetime.now(TZ)
    
    if now.hour < 8:
        start_date = now.date() - timedelta(days=1)
    else:
        start_date = now.date()
    
    for i in range(days):
        date = start_date + timedelta(days=i)
        
        if date < now.date():
            continue
            
        date_formatted = date.strftime('%d.%m.%Y')
        weekday = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Нд"][date.weekday()]
        
        if date.weekday() == 5:
            weekday_display = f"🟦 {weekday}"
        elif date.weekday() == 6:
            weekday_display = f"🟥 {weekday}"
        else:
            weekday_display = weekday
        
        if instructor_name:
            free_slots = get_available_time_slots(instructor_name, date_formatted)
            free_count = len(free_slots)
            
            if free_count > 0:
                formatted = f"{weekday_display} {date.strftime('%d.%m')} ({free_count})"
                dates.append(formatted)
        else:
            formatted = f"{weekday_display} {date.strftime('%d.%m.%Y')}"
            dates.append(formatted)
    
    return dates

def get_available_time_slots(instructor_name, date_str):
    """Отримати вільні часові слоти для інструктора"""
    try:
        instructor_data = get_instructor_by_name(instructor_name)
        if not instructor_data:
            return []
        
        instructor_id = instructor_data[0]
        
        date_obj = datetime.strptime(date_str, "%d.%m.%Y")
        now = datetime.now(TZ)
        is_today = date_obj.date() == now.date()
        
        all_slots = []
        
        if is_today:
            min_time = now + timedelta(hours=1)
            min_hour = min_time.hour
            
            if min_time.minute > 0:
                min_hour += 1
            
            start_hour = max(min_hour, WORK_HOURS_START)
        else:
            start_hour = WORK_HOURS_START
        
        if start_hour < WORK_HOURS_START:
            start_hour = WORK_HOURS_START
        
        hour = start_hour
        while hour < WORK_HOURS_END:
            all_slots.append(f"{hour:02d}:00")
            hour += 1
        
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT time, duration FROM lessons
                WHERE instructor_id = ? AND date = ? AND status = 'active'
            """, (instructor_id, date_str))
            booked = cursor.fetchall()
        
        blocked_hours = set()
        for booked_time, duration in booked:
            if ':' not in booked_time:
                continue
            
            start_h = int(booked_time.split(':')[0])
            
            if "1.5" in duration:
                hours_blocked = 2
            elif "2" in duration:
                hours_blocked = 2
            else:
                hours_blocked = 1
            
            for i in range(hours_blocked):
                blocked_hours.add(f"{start_h + i:02d}:00")
        
        from database import is_time_blocked
        date_formatted = date_obj.strftime("%Y-%m-%d")
        
        free_slots = [
            slot for slot in all_slots 
            if slot not in blocked_hours
            and not is_time_blocked(instructor_id, date_formatted, slot)
        ]
        
        return free_slots
        
    except Exception as e:
        logger.error(f"Помилка get_available_time_slots: {e}")
        return []

# ======================= VALIDATORS =======================
def validate_phone(phone):
    """Валідація українського номера"""
    clean = re.sub(r'[\s\-\(\)]', '', phone)
    patterns = [
        r'^(\+?38)?0\d{9}$',
        r'^\d{10}$'
    ]
    return any(re.match(p, clean) for p in patterns)

def validate_date_format(date_str):
    """Валідація формату дати"""
    try:
        datetime.strptime(date_str, "%d.%m.%Y")
        return True
    except ValueError:
        return False

def is_admin(user_id):
    """Перевірка чи користувач є адміном"""
    return user_id in ADMIN_ID

# ======================= START =======================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Головне меню"""
    user_id = update.message.from_user.id
    logger.info(f"🟢 START викликано! User: {user_id}, Args: {context.args}")
    
    if is_admin(user_id):
        try:
            with get_db() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id FROM instructors WHERE telegram_id = ?", (user_id,))
                is_instructor = cursor.fetchone() is not None
            
            if is_instructor:
                keyboard = [
                    [KeyboardButton("🚗 Автомат"), KeyboardButton("🚙 Механіка")],
                    [KeyboardButton("📅 Мій розклад")],
                    [KeyboardButton("⚙️ Управління графіком")],
                    [KeyboardButton("📊 Моя статистика")],
                    [KeyboardButton("❌ Історія скасувань")],
                    [KeyboardButton("⭐ Оцінити учня")],
                    [KeyboardButton("🔐 Панель адміна")]
                ]
                text = "Привіт! 👋 Я бот *Автоінструктор*.\n\n👨‍🏫 *Панель інструктора*\n🔐 *Панель адміністратора*\n\nОберіть дію:"
                context.user_data["state"] = "waiting_for_transmission"
            else:
                await show_admin_panel(update, context)
                return
            
            await update.message.reply_text(
                text,
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
                parse_mode="Markdown"
            )
            return
        except Exception as e:
            logger.error(f"Error checking admin status: {e}", exc_info=True)
    
    if context.args:
        command = context.args[0]
        logger.info(f"🔗 Deep link виявлено: {command} — доступ заблоковано")
        if command in ("register490", "register590"):
            await update.message.reply_text(
                "⛔ *Самостійна реєстрація заблокована*\n\n"
                "Для внесення вас в систему зверніться до адміністратора автошколи.\n\n"
                "📞 Контакт адміністратора: +380677499988/+380505475557",
                parse_mode="Markdown"
            )
            return
    
    context.user_data.clear()

    try:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM instructors WHERE telegram_id = ?", (user_id,))
            is_instructor = cursor.fetchone() is not None

        if is_instructor:
            keyboard = [
                [KeyboardButton("🚗 Автомат"), KeyboardButton("🚙 Механіка")],
                [KeyboardButton("📅 Мій розклад")],
                [KeyboardButton("⚙️ Управління графіком")],
                [KeyboardButton("📊 Моя статистика")],
                [KeyboardButton("❌ Історія скасувань")],
                [KeyboardButton("⭐ Оцінити учня")]
            ]
            text = "Привіт! 👋 Я бот *Автоінструктор*.\n\n👨‍🏫 *Панель інструктора*\n\nОберіть дію:"
            
            context.user_data["state"] = "waiting_for_transmission"
            
            await update.message.reply_text(
                text,
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
                parse_mode="Markdown"
            )
        else:
            student = get_student_by_telegram_id(user_id)
            
            if student:
                context.user_data["student_name"] = student[1]
                context.user_data["student_phone"] = student[2]
                context.user_data["student_tariff"] = student[3]
                
                keyboard = [
                    [KeyboardButton("🚀 Записатися на заняття")],
                    [KeyboardButton("📋 Мої записи")],
                    [KeyboardButton("❌ Скасувати запис")],
                    [KeyboardButton("📊 Моя статистика")]
                ]
                
                await update.message.reply_text(
                    f"Привіт, {student[1]}! 👋\n\n"
                    f"💰 Ваш тариф: {student[3]} грн/год\n\n"
                    f"Що бажаєте зробити?",
                    reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
                )
            else:
                await update.message.reply_text(
                    "⛔ *Ви не зареєстровані в системі*\n\n"
                    "Для отримання доступу зверніться до адміністратора автошколи — "
                    "він внесе вас в систему вручну.\n\n"
                    "Після реєстрації адміном напишіть /start ще раз.",
                    parse_mode="Markdown"
                )
        
    except Exception as e:
        logger.error(f"Error in start: {e}", exc_info=True)
        await update.message.reply_text("❌ Виникла помилка. Спробуйте /start")

# ======================= REGISTRATION COMMANDS =======================
async def register_490(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Реєстрація учня з тарифом 490 грн"""
    logger.info("🔵 register_490 викликано!")
    try:
        await register_student_with_tariff(update, context, 490)
    except Exception as e:
        logger.error(f"❌ Помилка в register_490: {e}", exc_info=True)

async def register_590(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Реєстрація учня з тарифом 590 грн"""
    await register_student_with_tariff(update, context, 590)

async def register_student_with_tariff(update: Update, context: ContextTypes.DEFAULT_TYPE, tariff: int):
    """Загальна функція реєстрації учня"""
    user = update.message.from_user
    user_id = user.id
    logger.info(f"🟡 register_student_with_tariff викликано! User: {user_id}, Tariff: {tariff}")
    
    student = get_student_by_telegram_id(user_id)
    
    if student:
        logger.info(f"✅ Учень вже зареєстрований: {student[1]}")
        await update.message.reply_text(
            f"✅ Ви вже зареєстровані!\n\n"
            f"👤 Ім'я: {student[1]}\n"
            f"💰 Тариф: {student[3]} грн/год\n\n"
            f"Використовуйте /start для запису на заняття."
        )
        return
    
    auto_name = f"{user.first_name or ''} {user.last_name or ''}".strip()
    auto_phone = user.username if user.username else ""
    logger.info(f"📝 Автозаповнення: name={auto_name}, username={auto_phone}")
    
    context.user_data["registration_tariff"] = tariff
    context.user_data["auto_name"] = auto_name
    context.user_data["state"] = "registration_name"
    
    keyboard = []
    if auto_name:
        keyboard.append([KeyboardButton(f"✅ {auto_name}")])
    keyboard.append([KeyboardButton("🔙 Скасувати")])
    
    logger.info(f"💬 Відправляю запит на введення імені")
    await update.message.reply_text(
        f"🎓 *Реєстрація учня*\n"
        f"💰 Тариф: *{tariff} грн/год*\n\n"
        f"Введіть ваше ім'я та прізвище:",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        parse_mode="Markdown"
    )

# ======================= HANDLE MESSAGE =======================
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Головний обробник"""
    text = update.message.text
    state = context.user_data.get("state", "")
    user_id = update.message.from_user.id
    
    logger.info(f"📥 Message: '{text}' | State: '{state}'")
    
    try:
        # === ОЦІНЮВАННЯ ІНСТРУКТОРА УЧНЕМ ===
        if text in ["⭐", "⭐⭐", "⭐⭐⭐", "⭐⭐⭐⭐", "⭐⭐⭐⭐⭐"]:
            lesson_data = context.bot_data.get(f"rating_lesson_{user_id}")
            
            if lesson_data:
                rating_map = {
                    "⭐": 1,
                    "⭐⭐": 2,
                    "⭐⭐⭐": 3,
                    "⭐⭐⭐⭐": 4,
                    "⭐⭐⭐⭐⭐": 5
                }
                rating = rating_map.get(text, 5)
                
                with get_db() as conn:
                    cursor = conn.cursor()
                    cursor.execute("""
                        UPDATE lessons
                        SET rating = ?
                        WHERE id = ?
                    """, (rating, lesson_data['lesson_id']))
                    conn.commit()
                
                context.bot_data[f"rating_feedback_{user_id}"] = {
                    'lesson_id': lesson_data['lesson_id'],
                    'instructor_name': lesson_data['instructor_name'],
                    'rating': rating
                }
                
                del context.bot_data[f"rating_lesson_{user_id}"]
                
                context.user_data["state"] = "rating_feedback"
                
                keyboard = [
                    [KeyboardButton("✍️ Написати коментар")],
                    [KeyboardButton("⏭️ Пропустити")]
                ]
                
                await update.message.reply_text(
                    f"✅ *Дякуємо за оцінку!*\n"
                    f"⭐ Оцінка: {rating}/5\n\n"
                    f"💬 Хочете залишити коментар?",
                    reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
                    parse_mode="Markdown"
                )
                
                logger.info(f"✅ Учень {user_id} оцінив урок {lesson_data['lesson_id']}: {rating}/5")
                return
            else:
                logger.warning(f"⚠️ Учень {user_id} надіслав оцінку але немає lesson_data")
        
        if text == "⏭️ Пропустити" and f"rating_lesson_{user_id}" in context.bot_data:
            lesson_data = context.bot_data.get(f"rating_lesson_{user_id}")
            del context.bot_data[f"rating_lesson_{user_id}"]
            
            await update.message.reply_text(
                f"✅ Дякуємо!\n\n"
                f"📅 {lesson_data['date']} {lesson_data['time']}\n"
                f"👨‍🏫 {lesson_data['instructor_name']}"
            )
            
            logger.info(f"⏭️ Учень {user_id} пропустив оцінювання уроку {lesson_data['lesson_id']}")
            await start(update, context)
            return
        
        if state == "rating_feedback":
            feedback_data = context.bot_data.get(f"rating_feedback_{user_id}")
            
            if text == "✍️ Написати коментар":
                context.user_data["state"] = "rating_feedback_input"
                
                keyboard = [[KeyboardButton("⏭️ Пропустити")]]
                
                await update.message.reply_text(
                    "💬 Введіть ваш коментар:",
                    reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
                )
                return
            
            elif text == "⏭️ Пропустити":
                if feedback_data:
                    del context.bot_data[f"rating_feedback_{user_id}"]
                    context.user_data.clear()
                    
                    await update.message.reply_text(
                        f"✅ *Дякуємо за відгук!*\n\n"
                        f"👨‍🏫 {feedback_data['instructor_name']}\n"
                        f"⭐ Оцінка: {feedback_data['rating']}/5",
                        parse_mode="Markdown"
                    )
                    
                    logger.info(f"⏭️ Учень {user_id} пропустив коментар для уроку {feedback_data['lesson_id']}")
                    await start(update, context)
                    return
        
        if state == "rating_feedback_input":
            feedback_data = context.bot_data.get(f"rating_feedback_{user_id}")
            
            if text == "⏭️ Пропустити":
                if feedback_data:
                    del context.bot_data[f"rating_feedback_{user_id}"]
                    context.user_data.clear()
                    
                    await update.message.reply_text(
                        f"✅ *Дякуємо за відгук!*\n\n"
                        f"👨‍🏫 {feedback_data['instructor_name']}\n"
                        f"⭐ Оцінка: {feedback_data['rating']}/5",
                        parse_mode="Markdown"
                    )
                    
                    logger.info(f"⏭️ Учень {user_id} пропустив коментар для уроку {feedback_data['lesson_id']}")
                    await start(update, context)
                    return
            else:
                feedback_text = text
                
                with get_db() as conn:
                    cursor = conn.cursor()
                    cursor.execute("""
                        UPDATE lessons
                        SET feedback = ?
                        WHERE id = ?
                    """, (feedback_text, feedback_data['lesson_id']))
                    conn.commit()
                
                del context.bot_data[f"rating_feedback_{user_id}"]
                context.user_data.clear()
                
                await update.message.reply_text(
                    f"✅ *Дякуємо за відгук!*\n\n"
                    f"👨‍🏫 {feedback_data['instructor_name']}\n"
                    f"⭐ Оцінка: {feedback_data['rating']}/5\n"
                    f"💬 \"{feedback_text}\"",
                    parse_mode="Markdown"
                )
                
                logger.info(f"✅ Учень {user_id} залишив коментар для уроку {feedback_data['lesson_id']}")
                await start(update, context)
                return
        
        # === РЕЄСТРАЦІЯ УЧНЯ ===
        if state == "registration_name":
            if text == "🔙 Скасувати":
                await update.message.reply_text("❌ Реєстрацію скасовано.")
                return
            
            context.user_data["student_name"] = text
            context.user_data["state"] = "registration_phone"
            
            keyboard = [[KeyboardButton("📱 Надати номер", request_contact=True)]]
            keyboard.append([KeyboardButton("🔙 Скасувати")])
            
            await update.message.reply_text(
                "📱 Тепер надайте ваш номер телефону:\n"
                "(натисніть кнопку нижче або введіть вручну)",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True)
            )
            return
        
        if state == "registration_phone":
            if text == "🔙 Скасувати":
                await update.message.reply_text("❌ Реєстрацію скасовано.")
                return
            
            phone = None
            if update.message.contact:
                phone = update.message.contact.phone_number
            elif validate_phone(text):
                phone = text
            else:
                await update.message.reply_text("⚠️ Невірний формат номера. Спробуйте ще раз:")
                return
            
            user_id = update.message.from_user.id
            name = context.user_data["student_name"]
            tariff = context.user_data["registration_tariff"]
            
            if register_student(name, phone, user_id, tariff, f"link_{tariff}"):
                keyboard = [
                    [KeyboardButton("🚀 Записатися на заняття")],
                    [KeyboardButton("📋 Мої записи")]
                ]
                
                await update.message.reply_text(
                    f"✅ *Реєстрацію завершено!*\n\n"
                    f"👤 Ім'я: {name}\n"
                    f"📱 Телефон: {phone}\n"
                    f"💰 Ваш тариф: *{tariff} грн/год* (фіксований)\n\n"
                    f"ℹ️ Тариф закріплений за вами і не змінюється.\n\n"
                    f"Натисніть кнопку нижче, щоб записатися на заняття:",
                    reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
                    parse_mode="Markdown"
                )
            else:
                await update.message.reply_text("❌ Помилка реєстрації. Спробуйте пізніше.")
            
            context.user_data.clear()
            return

        # === ПАНЕЛЬ АДМІНА ===
        if text == "🔐 Панель адміна":
            if not is_admin(update.message.from_user.id):
                await update.message.reply_text("❌ У вас немає доступу.")
                return
            await show_admin_panel(update, context)
            return
        
        if text == "📥 Експорт в Excel":
            await show_export_period_menu(update, context)
            return
        
        if state == "admin_panel":
            if update.message.text == "✏️ Управління записами":
                await handle_admin_manage_bookings(update, context)
                return
            if update.message.text == "➕ Додати учня":
                await admin_add_student_start(update, context)
                return
            await handle_admin_report(update, context)
            return
        
        # === ДОДАВАННЯ УЧНЯ АДМІНОМ ===
        if state == "admin_add_student_name":
            await handle_admin_add_student_name(update, context)
            return
        
        if state == "admin_add_student_phone":
            await handle_admin_add_student_phone(update, context)
            return
        
        if state == "admin_add_student_tariff":
            await handle_admin_add_student_tariff(update, context)
            return
        
        if state == "admin_add_student_tgid":
            await handle_admin_add_student_tgid(update, context)
            return
        
        if state == "admin_manage_bookings":
            await handle_admin_manage_bookings(update, context)
            return
        
        if state == "admin_cancel_select_date":
            await handle_admin_cancel_select_date(update, context)
            return
        
        if state == "admin_cancel_select_instructor":
            await handle_admin_cancel_select_instructor(update, context)
            return
        
        if state == "admin_cancel_select_lesson":
            await handle_admin_cancel_select_lesson(update, context)
            return
        
        # === РУЧНИЙ ЗАПИС УЧНЯ АДМІНОМ ===
        if state == "admin_manual_enter_phone":
            await handle_admin_manual_enter_phone(update, context)
            return
        
        if state == "admin_manual_confirm_student":
            await handle_admin_manual_confirm_student(update, context)
            return
        
        if state == "admin_manual_enter_name":
            await handle_admin_manual_enter_name(update, context)
            return
        
        if state == "admin_manual_select_tariff":
            await handle_admin_manual_select_tariff(update, context)
            return
        
        if state == "admin_manual_select_transmission":
            await handle_admin_manual_select_transmission(update, context)
            return
        
        if state == "admin_manual_select_instructor":
            await handle_admin_manual_select_instructor(update, context)
            return
        
        if state == "admin_manual_select_date":
            await handle_admin_manual_select_date(update, context)
            return
        
        if state == "admin_manual_select_time":
            await handle_admin_manual_select_time(update, context)
            return
        
        if state == "admin_manual_select_duration":
            await handle_admin_manual_select_duration(update, context)
            return
        
        if state == "admin_manual_confirm":
            await handle_admin_manual_confirm(update, context)
            return
        
        if state == "admin_select_instructor_report":
            await handle_instructor_report_select(update, context)
            return
        if state == "admin_instructor_report_period":
            await handle_instructor_report_period(update, context)
            return
        if state == "admin_instructor_custom_period":
            await handle_instructor_custom_period(update, context)
            return
        if state == "admin_report_period":
            await handle_admin_report(update, context)
            return
        
        # === ЕКСПОРТ З ВИБОРОМ ПЕРІОДУ ===
        if state == "export_period":
            await handle_export_period_choice(update, context)
            return
        
        if state == "export_custom_period":
            await handle_export_custom_period(update, context)
            return

        # === МЕНЮ ІНСТРУКТОРА ===
        if text == "🔙 Назад":
            with get_db() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id FROM instructors WHERE telegram_id = ?", (user_id,))
                is_instructor = cursor.fetchone() is not None
            
            if is_instructor:
                await start(update, context)
                return
        
        if text == "📅 Мій розклад":
            await show_instructor_schedule(update, context)
            return
        elif text == "⚙️ Управління графіком":
            await manage_schedule(update, context)
            return
        elif text == "📊 Моя статистика":
            with get_db() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id FROM instructors WHERE telegram_id = ?", (user_id,))
                is_instructor = cursor.fetchone() is not None
            
            if is_instructor:
                await show_instructor_stats_menu(update, context)
            else:
                await show_student_statistics(update, context)
            return
        elif text == "❌ Історія скасувань":
            await show_cancellation_history(update, context)
            return
        elif text == "⭐ Оцінити учня":
            await rate_student_menu(update, context)
            return

        # === СТАТИСТИКА ЗА ПЕРІОД ===
        if state == "stats_period":
            await handle_stats_period(update, context)
            return
        
        if state == "stats_custom_period":
            await handle_stats_custom_period(update, context)
            return
        
        # === ОЦІНЮВАННЯ УЧНЯ ===
        if state in ["rating_select_lesson", "rating_give_score", "rating_give_feedback"]:
            await handle_rating_flow(update, context)
            return

        # === КОРИГУВАННЯ ГРАФІКУ ===
        if state in ["edit_schedule_select", "edit_schedule_date", "edit_schedule_time"]:
            await handle_edit_schedule(update, context)
            return

        # === УПРАВЛІННЯ ГРАФІКОМ ===
        if state in ["schedule_menu", "block_choose_date", "block_choose_time_start", 
                     "block_choose_time_end", "block_choose_reason", "unblock_choose_date", "waiting_unblock"]:
            await handle_schedule_management(update, context)
            return

        # === МЕНЮ СТУДЕНТА ===
        if text == "🚀 Записатися на заняття":
            keyboard = [
                [KeyboardButton("🚗 Автомат"), KeyboardButton("🚙 Механіка")]
            ]
            context.user_data["state"] = "waiting_for_transmission"
            
            await update.message.reply_text(
                "🚗 Оберіть тип коробки передач:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
            return
        
        if text == "📖 Мої записи" or text == "📋 Мої записи":
            await show_student_lessons(update, context)
            return
        
        if text == "❌ Скасувати запис":
            await show_lessons_to_cancel(update, context)
            return
        
        # === СКАСУВАННЯ ЗАПИСУ ===
        if state == "cancel_lesson_select":
            await handle_cancel_lesson(update, context)
            return
        
        if state == "cancel_lesson_confirm":
            await handle_cancel_confirmation(update, context)
            return
        
        # === ПІДТВЕРДЖЕННЯ ===
        if state == "waiting_for_confirmation":
            if text == "✅ Підтвердити":
                await save_lesson(update, context)
                return
            elif text in ["💬 Додати коментар", "✏️ Змінити коментар"]:
                context.user_data["state"] = "waiting_for_booking_comment"
                
                keyboard = [
                    [KeyboardButton("⏭️ Пропустити")],
                    [KeyboardButton("🔙 Назад")]
                ]
                
                await update.message.reply_text(
                    "💬 *Введіть коментар для інструктора:*\n\n"
                    "_Наприклад:_\n"
                    "• \"Перше заняття\"\n"
                    "• \"буду чекати в Тисмениці\"\n"
                    "• \"практичний іспиту\"",
                    reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
                    parse_mode="Markdown"
                )
                return
            elif text == "🔙 Скасувати":
                await update.message.reply_text("❌ Запис скасовано.")
                await start(update, context)
                return
        
        # === ВВЕДЕННЯ КОМЕНТАРЯ ===
        if state == "waiting_for_booking_comment":
            if text == "🔙 Назад":
                context.user_data["state"] = "waiting_for_confirmation"
                await show_booking_confirmation(update, context)
                return
            elif text == "⏭️ Пропустити":
                context.user_data["booking_comment"] = ""
                await show_booking_confirmation(update, context)
                return
            else:
                context.user_data["booking_comment"] = text
                
                await update.message.reply_text(
                    f"✅ Коментар збережено!\n\n"
                    f"💬 \"{text}\""
                )
                
                await show_booking_confirmation(update, context)
                return
        
        # === МЕНЮ ВИБОРУ ПЕРІОДУ РОЗКЛАДУ ===
        if state == "instructor_schedule_menu":
            now = datetime.now(TZ)
            today = now.date()
            
            if text == "📅 На сьогодні":
                await show_instructor_schedule_period(update, context, date_from=today, date_to=today)
                return
            elif text == "📅 На завтра":
                tomorrow = today + timedelta(days=1)
                await show_instructor_schedule_period(update, context, date_from=tomorrow, date_to=tomorrow)
                return
            elif text == "📅 На тиждень":
                week_end = today + timedelta(days=6)
                await show_instructor_schedule_period(update, context, date_from=today, date_to=week_end)
                return
            elif text == "📅 Свій період":
                context.user_data["state"] = "instructor_schedule_custom_period"
                keyboard = [[KeyboardButton("🔙 Назад")]]
                await update.message.reply_text(
                    "📅 Введіть період у форматі:\n"
                    "*ДД.ММ.РРРР - ДД.ММ.РРРР*\n\n"
                    "Наприклад: 01.03.2026 - 15.03.2026",
                    reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
                    parse_mode="Markdown"
                )
                return
            elif text == "🔙 Назад":
                await start(update, context)
                return
        
        # === ВВЕДЕННЯ СВОГО ПЕРІОДУ ===
        if state == "instructor_schedule_custom_period":
            if text == "🔙 Назад":
                await show_instructor_schedule(update, context)
                return
            
            try:
                parts = text.split('-')
                if len(parts) != 2:
                    raise ValueError("Неправильний формат")
                
                date_from_str = parts[0].strip()
                date_to_str = parts[1].strip()
                
                date_from = datetime.strptime(date_from_str, "%d.%m.%Y").date()
                date_to = datetime.strptime(date_to_str, "%d.%m.%Y").date()
                
                if date_from > date_to:
                    await update.message.reply_text("❌ Дата початку не може бути пізніше дати кінця.")
                    return
                
                await show_instructor_schedule_period(update, context, date_from=date_from, date_to=date_to)
                return
                
            except Exception as e:
                await update.message.reply_text(
                    "❌ Неправильний формат дати.\n\n"
                    "Використовуйте формат: *ДД.ММ.РРРР - ДД.ММ.РРРР*\n"
                    "Наприклад: 01.03.2026 - 15.03.2026",
                    parse_mode="Markdown"
                )
                return
        
        # === ВИБІР КОРОБКИ ===
        if state == "waiting_for_transmission":
            if text == "👨‍🏫 Обрати іншого інструктора":
                transmission = context.user_data.get("transmission")
                if transmission:
                    context.user_data["state"] = "waiting_for_instructor"
                    instructors = get_instructors_by_transmission(transmission)
                    if not instructors:
                        await update.message.reply_text("😔 Немає інструкторів для цього типу.")
                        return

                    keyboard = []
                    for instructor in instructors:
                        rating = get_instructor_rating(instructor)
                        if rating > 0:
                            stars = "⭐" * int(rating)
                            keyboard.append([f"{instructor} {stars} ({rating:.1f})"])
                        else:
                            keyboard.append([f"{instructor} 🆕"])
                    
                    keyboard.append([KeyboardButton("🔙 Назад")])
                    
                    await update.message.reply_text(
                        "👨‍🏫 Оберіть інструктора:",
                        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
                    )
                    return
            
            if text not in ["🚗 Автомат", "🚙 Механіка"]:
                await update.message.reply_text("⚠️ Оберіть коробку передач із меню.")
                return
            
            transmission = "Автомат" if text == "🚗 Автомат" else "Механіка"
            context.user_data["transmission"] = transmission
            context.user_data["state"] = "waiting_for_instructor"

            instructors = get_instructors_by_transmission(transmission)
            if not instructors:
                await update.message.reply_text("😔 Немає інструкторів для цього типу.")
                return

            keyboard = []
            for instructor in instructors:
                rating = get_instructor_rating(instructor)
                if rating > 0:
                    stars = "⭐" * int(rating)
                    keyboard.append([f"{instructor} {stars} ({rating:.1f})"])
                else:
                    keyboard.append([f"{instructor} 🆕"])
            
            keyboard.append([KeyboardButton("🔙 Назад")])
            
            await update.message.reply_text(
                "👨‍🏫 Оберіть інструктора:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
            return
        
        # === ВИБІР ІНСТРУКТОРА ===
        if state == "waiting_for_instructor":
            logger.info(f"👨‍🏫 Обробка вибору інструктора: {text}")
            
            if text == "🔙 Назад":
                await start(update, context)
                return
            
            instructor_name = text.split(" ⭐")[0].split(" 🆕")[0]
            context.user_data["instructor"] = instructor_name
            context.user_data["state"] = "waiting_for_date"
            
            dates = get_next_dates(14, instructor_name)
            
            if not dates:
                keyboard = [
                    [KeyboardButton("👨‍🏫 Обрати іншого інструктора")],
                    [KeyboardButton("🔙 Назад")]
                ]
                await update.message.reply_text(
                    f"😔 У інструктора {instructor_name} всі години зайняті на найближчі 14 днів\n\n"
                    f"💡 Що робити:\n"
                    f"• Оберіть іншого інструктора - у них можуть бути вільні години\n"
                    f"• Зайдіть завтра після 8:00 - бот оновиться і з'являться нові дати для запису\n\n"
                    f"📅 Вільні години оновлюються щодня о 8:00 ранку",
                    reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
                )
                context.user_data["state"] = "waiting_for_transmission"
                return
            
            keyboard = []
            for i in range(0, len(dates), 2):
                row = [KeyboardButton(dates[i])]
                if i + 1 < len(dates):
                    row.append(KeyboardButton(dates[i + 1]))
                keyboard.append(row)
            
            keyboard.append([KeyboardButton("🔙 Назад")])
            
            await update.message.reply_text(
                f"📅 Оберіть дату заняття:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
            return
        
        # === ВИБІР ДАТИ ===
        if state == "waiting_for_date":
            logger.info(f"🔵 Обробка дати: {text}")
            
            if text == "🔙 Назад":
                transmission = context.user_data.get("transmission")
                instructors = get_instructors_by_transmission(transmission)
                
                context.user_data["state"] = "waiting_for_instructor"
                
                keyboard = []
                for instructor in instructors:
                    rating = get_instructor_rating(instructor)
                    if rating > 0:
                        stars = "⭐" * int(rating)
                        keyboard.append([f"{instructor} {stars} ({rating:.1f})"])
                    else:
                        keyboard.append([f"{instructor} 🆕"])
                
                keyboard.append([KeyboardButton("🔙 Назад")])
                
                await update.message.reply_text(
                    "👨‍🏫 Оберіть інструктора:",
                    reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
                )
                return
            
            valid_date_markers = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Нд", "🟦", "🟥"]
            if not any(marker in text for marker in valid_date_markers):
                logger.warning(f"⚠️ Спроба ввести дату вручну: {text}")
                await update.message.reply_text(
                    "⚠️ Будь ласка, оберіть дату з кнопок нижче.\n\n"
                    "Якщо потрібної дати немає у списку - зверніться до адміністратора або оберіть іншого інструктора."
                )
                return
            
            date_parts = text.split()
            
            if len(date_parts) >= 3 and date_parts[0] in ["🟦", "🟥"]:
                date_candidate = date_parts[2].replace("(", "").replace(")", "")
            elif len(date_parts) >= 2:
                date_candidate = date_parts[1]
            else:
                date_str = text
                date_candidate = None
            
            if date_candidate:
                if date_candidate.count('.') == 1:
                    current_year = datetime.now().year
                    date_str = f"{date_candidate}.{current_year}"
                else:
                    date_str = date_candidate
            
            logger.info(f"📆 Витягнута дата: {date_str}")
            
            if not validate_date_format(date_str):
                logger.warning(f"⚠️ Невірний формат дати: {date_str}")
                await update.message.reply_text(
                    "⚠️ Невірний формат дати. Оберіть дату з меню."
                )
                return
            
            date_obj = datetime.strptime(date_str, "%d.%m.%Y")
            today = datetime.now(TZ).date()
            if date_obj.date() < today:
                logger.warning(f"⚠️ Минула дата: {date_str} (сьогодні: {today})")
                await update.message.reply_text("⚠️ Неможливо записатися на минулу дату.")
                return
            
            context.user_data["date"] = date_str
            instructor = context.user_data["instructor"]
            
            free_slots = get_available_time_slots(instructor, date_str)
            
            if not free_slots:
                await update.message.reply_text(
                    "😔 На цю дату немає вільних місць.\n"
                    "Оберіть іншу дату:"
                )
                return
            
            context.user_data["state"] = "waiting_for_time"
            
            keyboard = []
            for i in range(0, len(free_slots), 3):
                row = []
                for j in range(3):
                    if i + j < len(free_slots):
                        row.append(KeyboardButton(free_slots[i + j]))
                keyboard.append(row)
            
            keyboard.append([KeyboardButton("🔙 Назад")])
            
            await update.message.reply_text(
                "🕐 Оберіть час заняття:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
            return
        
        # === ВИБІР ЧАСУ ===
        if state == "waiting_for_time":
            if text == "🔙 Назад":
                context.user_data["state"] = "waiting_for_date"
                await update.message.reply_text("📅 Введіть іншу дату (ДД.ММ.РРРР):")
                return
            
            if not re.match(r'^([0-1][0-9]|2[0-3]):[0-5][0-9]$', text):
                await update.message.reply_text(
                    "⚠️ Будь ласка, оберіть час з кнопок нижче.\n\n"
                    "Якщо потрібного часу немає - оберіть іншу дату або інструктора."
                )
                return
            
            instructor = context.user_data.get("instructor")
            date = context.user_data.get("date")
            free_slots = get_available_time_slots(instructor, date)
            
            if text not in free_slots:
                await update.message.reply_text(
                    "⚠️ Цей час недоступний. Будь ласка, оберіть час з доступних варіантів."
                )
                return
            
            context.user_data["time"] = text
            context.user_data["state"] = "waiting_for_duration"
            
            keyboard = [
                [KeyboardButton("1 година")],
                [KeyboardButton("2 години")],
                [KeyboardButton("🔙 Назад")]
            ]
            
            await update.message.reply_text(
                "⏱ Оберіть тривалість заняття:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
            return
        
        # === ВИБІР ТРИВАЛОСТІ ===
        if state == "waiting_for_duration":
            if text == "🔙 Назад":
                instructor = context.user_data["instructor"]
                date = context.user_data["date"]
                free_slots = get_available_time_slots(instructor, date)
                
                context.user_data["state"] = "waiting_for_time"
                
                keyboard = []
                for i in range(0, len(free_slots), 3):
                    row = []
                    for j in range(3):
                        if i + j < len(free_slots):
                            row.append(KeyboardButton(free_slots[i + j]))
                    keyboard.append(row)
                
                keyboard.append([KeyboardButton("🔙 Назад")])
                
                await update.message.reply_text(
                    "🕐 Оберіть час заняття:",
                    reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
                )
                return
            
            if text not in ["1 година", "2 години"]:
                await update.message.reply_text("⚠️ Оберіть тривалість із меню.")
                return
            
            if text == "2 години":
                selected_time = context.user_data["time"]
                instructor = context.user_data["instructor"]
                date = context.user_data["date"]
                
                selected_hour = int(selected_time.split(':')[0])
                next_hour = f"{selected_hour + 1:02d}:00"
                
                free_slots = get_available_time_slots(instructor, date)
                
                if next_hour not in free_slots and next_hour != f"{WORK_HOURS_END:02d}:00":
                    await update.message.reply_text(
                        "⚠️ Наступна година зайнята. Оберіть інший час або 1 годину."
                    )
                    return
            
            context.user_data["duration"] = text
            
            user = update.message.from_user
            student = get_student_by_telegram_id(user.id)
            
            if student:
                student_id = student[0]
                selected_date = context.user_data["date"]
                selected_duration = text
                
                duration_hours = 2 if "2" in selected_duration else 1
                
                lesson_date = datetime.strptime(selected_date, "%d.%m.%Y")
                
                week_start = lesson_date - timedelta(days=lesson_date.weekday())
                week_end = week_start + timedelta(days=6)
                
                with get_db() as conn:
                    cursor = conn.cursor()
                    
                    cursor.execute("""
                        SELECT date, duration
                        FROM lessons
                        WHERE student_telegram_id = ? AND status = 'active'
                    """, (user.id,))
                    all_lessons = cursor.fetchall()
                    
                    hours_this_week = 0
                    for lesson_date_str, lesson_duration in all_lessons:
                        try:
                            ld = datetime.strptime(lesson_date_str, "%d.%m.%Y")
                            if week_start <= ld <= week_end:
                                if "2" in str(lesson_duration):
                                    hours_this_week += 2
                                elif "1.5" in str(lesson_duration):
                                    hours_this_week += 1.5
                                else:
                                    hours_this_week += 1
                        except:
                            pass
                
                if hours_this_week + duration_hours > 6:
                    remaining = 6 - hours_this_week
                    await update.message.reply_text(
                        f"⚠️ Перевищено ліміт!\n\n"
                        f"На цей тиждень у вас вже заброньовано {hours_this_week:.1f} год.\n"
                        f"Ліміт: 6 годин на тиждень\n"
                        f"Доступно: {remaining:.1f} год.\n\n"
                        f"Оберіть інший тиждень або зменшіть тривалість."
                    )
                    return
                
                context.user_data["student_name"] = student[1]
                context.user_data["student_phone"] = student[2]
                context.user_data["student_tariff"] = student[3]
                
                await show_booking_confirmation(update, context)
            else:
                await update.message.reply_text(
                    "⚠️ *Помилка!*\n\n"
                    "Для запису потрібна реєстрація через спеціальне посилання.\n"
                    "Зверніться до адміністратора.",
                    parse_mode="Markdown"
                )
                await start(update, context)
            return
        
        # === ІМ'Я СТУДЕНТА ===
        if state == "waiting_for_name":
            if text == "🔙 Назад":
                if "duration" not in context.user_data:
                    await start(update, context)
                    return
                
                context.user_data["state"] = "waiting_for_duration"
                keyboard = [
                    [KeyboardButton("1 година")],
                    [KeyboardButton("2 години")],
                    [KeyboardButton("🔙 Назад")]
                ]
                await update.message.reply_text(
                    "⏱ Оберіть тривалість:",
                    reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
                )
                return
            
            if text.startswith("✅ "):
                text = text[2:]
            
            context.user_data["student_name"] = text
            context.user_data["state"] = "waiting_for_phone"
            
            keyboard = [[KeyboardButton("📱 Надати номер", request_contact=True)]]
            keyboard.append([KeyboardButton("🔙 Назад")])
            
            await update.message.reply_text(
                "📱 Введіть номер телефону або натисніть кнопку нижче:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True)
            )
            return
        
        # === ТЕЛЕФОН СТУДЕНТА ===
        if state == "waiting_for_phone":
            if text == "🔙 Назад":
                user = update.message.from_user
                auto_name = f"{user.first_name or ''} {user.last_name or ''}".strip()
                keyboard = []
                if auto_name:
                    keyboard.append([KeyboardButton(f"✅ {auto_name}")])
                keyboard.append([KeyboardButton("🔙 Назад")])
                
                context.user_data["state"] = "waiting_for_name"
                await update.message.reply_text(
                    "👤 Введіть ваше ім'я:",
                    reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
                )
                return
            
            phone = None
            if update.message.contact:
                phone = update.message.contact.phone_number
            elif validate_phone(text):
                phone = text
            else:
                await update.message.reply_text("⚠️ Невірний формат номера. Спробуйте ще раз:")
                return
            
            context.user_data["student_phone"] = phone
            
            if "duration" not in context.user_data:
                context.user_data["state"] = "waiting_for_transmission"
                
                keyboard = [
                    [KeyboardButton("🚗 Автомат"), KeyboardButton("🚙 Механіка")]
                ]
                
                await update.message.reply_text(
                    "✅ Дякую! Тепер оберіть тип коробки передач:",
                    reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
                )
            else:
                await show_booking_confirmation(update, context)
            return
        
    except Exception as e:
        logger.error(f"Error in handle_message: {e}", exc_info=True)
        await update.message.reply_text("❌ Виникла помилка. Спробуйте /start")

async def show_booking_confirmation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показати підтвердження бронювання"""
    instructor = context.user_data["instructor"]
    date = context.user_data["date"]
    time = context.user_data["time"]
    duration = context.user_data["duration"]
    name = context.user_data.get("student_name", "")
    phone = context.user_data.get("student_phone", "")
    student_tariff = context.user_data.get("student_tariff", 0)
    booking_comment = context.user_data.get("booking_comment", "")
    
    if student_tariff > 0:
        if "2" in duration:
            price = student_tariff * 2
        else:
            price = student_tariff
    else:
        price = PRICES.get(duration, 420)
    
    context.user_data["state"] = "waiting_for_confirmation"
    
    text = (
        f"📋 *Підтвердження запису*\n\n"
        f"👨‍🏫 Інструктор: {instructor}\n"
        f"📅 Дата: {date}\n"
        f"🕐 Час: {time}\n"
        f"⏱ Тривалість: {duration}\n"
        f"💰 Вартість: {price:.0f} грн\n"
    )
    
    if booking_comment:
        text += f"\n💬 Коментар:\n\"{booking_comment}\"\n"
    
    text += "\nВсе вірно?"
    
    keyboard = [
        [KeyboardButton("✅ Підтвердити")]
    ]
    
    if booking_comment:
        keyboard.append([KeyboardButton("✏️ Змінити коментар")])
    else:
        keyboard.append([KeyboardButton("💬 Додати коментар")])
    
    keyboard.append([KeyboardButton("🔙 Скасувати")])
    
    await update.message.reply_text(
        text,
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        parse_mode="Markdown"
    )

# ======================= INSTRUCTOR FUNCTIONS =======================
async def show_instructor_schedule(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показати меню вибору періоду для розкладу"""
    user_id = update.message.from_user.id
    
    try:
        instructor_data = get_instructor_by_telegram_id(user_id)
        if not instructor_data:
            await update.message.reply_text("❌ Ви не зареєстровані як інструктор.")
            return
        
        keyboard = [
            [KeyboardButton("📅 На сьогодні"), KeyboardButton("📅 На завтра")],
            [KeyboardButton("📅 На тиждень")],
            [KeyboardButton("📅 Свій період")],
            [KeyboardButton("🔙 Назад")]
        ]
        
        context.user_data["state"] = "instructor_schedule_menu"
        
        await update.message.reply_text(
            "📅 *Мій розклад*\n\nОберіть період для перегляду:",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
            parse_mode="Markdown"
        )
        
    except Exception as e:
        logger.error(f"Error in show_instructor_schedule: {e}", exc_info=True)
        await update.message.reply_text("❌ Помилка завантаження меню.")


async def show_instructor_schedule_period(update: Update, context: ContextTypes.DEFAULT_TYPE, date_from=None, date_to=None):
    """Показати розклад інструктора за вказаний період"""
    user_id = update.message.from_user.id
    
    try:
        instructor_data = get_instructor_by_telegram_id(user_id)
        if not instructor_data:
            await update.message.reply_text("❌ Ви не зареєстровані як інструктор.")
            return
        
        instructor_id, instructor_name = instructor_data
        
        now = datetime.now(TZ)
        
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT date, time, duration, student_name, student_phone, status, booking_comment
                FROM lessons
                WHERE instructor_id = ? 
                AND status = 'active'
                ORDER BY date, time
            """, (instructor_id,))
            
            all_lessons = cursor.fetchall()
        
        lessons = []
        for date, time, duration, student_name, student_phone, status, booking_comment in all_lessons:
            try:
                lesson_datetime = datetime.strptime(f"{date} {time}", "%d.%m.%Y %H:%M")
                lesson_datetime = TZ.localize(lesson_datetime)
                lesson_date = lesson_datetime.date()
                
                if lesson_datetime >= now:
                    if date_from and lesson_date < date_from:
                        continue
                    if date_to and lesson_date > date_to:
                        continue
                    
                    lessons.append((lesson_datetime, date, time, duration, student_name, student_phone, status, booking_comment))
            except:
                pass
        
        lessons.sort(key=lambda x: x[0])
        lessons = [(d, t, dur, sn, sp, st, bc) for (_, d, t, dur, sn, sp, st, bc) in lessons[:100]]
        
        if not lessons:
            keyboard = [[KeyboardButton("🔙 Назад")]]
            await update.message.reply_text(
                "📋 У вас поки немає запланованих занять за цей період.",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
            return
        
        if date_from and date_to:
            if date_from == date_to:
                period_text = f"на {date_from.strftime('%d.%m.%Y')}"
            else:
                period_text = f"з {date_from.strftime('%d.%m.%Y')} по {date_to.strftime('%d.%m.%Y')}"
        else:
            period_text = ""
        
        messages = []
        current_message = f"📅 *Ваш розклад {period_text}:*\n\n"
        current_date = None
        
        for date, time, duration, student_name, student_phone, status, booking_comment in lessons:
            lesson_text = ""
            if date != current_date:
                lesson_text += f"\n📆 *{date}*\n"
                current_date = date
            
            lesson_text += f"🕐 {time} ({duration})\n"
            lesson_text += f"👤 {student_name}\n"
            if student_phone:
                lesson_text += f"📱 {student_phone}\n"
            if booking_comment:
                lesson_text += f"💬 \"{booking_comment}\"\n"
            lesson_text += "\n"
            
            if len(current_message + lesson_text) > 3000:
                messages.append(current_message)
                current_message = lesson_text
            else:
                current_message += lesson_text
        
        if current_message:
            messages.append(current_message)
        
        for i, msg in enumerate(messages):
            if i == len(messages) - 1:
                keyboard = [[KeyboardButton("🔙 Назад")]]
                await update.message.reply_text(
                    msg,
                    reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
                    parse_mode="Markdown"
                )
            else:
                await update.message.reply_text(msg, parse_mode="Markdown")
        
    except Exception as e:
        logger.error(f"Error in show_instructor_schedule: {e}", exc_info=True)
        await update.message.reply_text("❌ Помилка завантаження розкладу.")

async def show_instructor_stats_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Меню вибору періоду для статистики"""
    keyboard = [
        [KeyboardButton("📊 За сьогодні")],
        [KeyboardButton("📊 За тиждень")],
        [KeyboardButton("📊 За місяць")],
        [KeyboardButton("📊 Свій період")],
        [KeyboardButton("🔙 Назад")]
    ]
    
    await update.message.reply_text(
        "📊 Статистика\n\nОберіть період:",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    )
    
    context.user_data["state"] = "stats_period"

async def handle_stats_period(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обробка вибору періоду статистики"""
    text = update.message.text
    user_id = update.message.from_user.id
    
    if text == "🔙 Назад":
        await start(update, context)
        return
    
    instructor_data = get_instructor_by_telegram_id(user_id)
    if not instructor_data:
        await update.message.reply_text("❌ Помилка.")
        return
    
    instructor_id, instructor_name = instructor_data
    
    today = datetime.now().date()
    
    if text == "📊 За сьогодні":
        date_from = today.strftime("%d.%m.%Y")
        date_to = today.strftime("%d.%m.%Y")
        period_text = "сьогодні"
    elif text == "📊 За тиждень":
        date_from = (today - timedelta(days=7)).strftime("%d.%m.%Y")
        date_to = today.strftime("%d.%m.%Y")
        period_text = "за тиждень"
    elif text == "📊 За місяць":
        date_from = (today - timedelta(days=30)).strftime("%d.%m.%Y")
        date_to = today.strftime("%d.%m.%Y")
        period_text = "за місяць"
    elif text == "📊 Свій період":
        context.user_data["state"] = "stats_custom_period"
        await update.message.reply_text(
            "📅 Введіть період у форматі:\n"
            "ДД.ММ.РРРР - ДД.ММ.РРРР\n\n"
            "Наприклад: 01.11.2024 - 30.11.2024"
        )
        return
    else:
        await update.message.reply_text("⚠️ Оберіть період із меню.")
        return
    
    await show_instructor_stats(update, context, instructor_id, date_from, date_to, period_text)

async def handle_stats_custom_period(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обробка введеного користувачем періоду"""
    text = update.message.text
    user_id = update.message.from_user.id
    
    if text == "🔙 Назад":
        await show_instructor_stats_menu(update, context)
        return
    
    try:
        import re
        match = re.match(r'(\d{2}\.\d{2}\.\d{4})\s*-\s*(\d{2}\.\d{2}\.\d{4})', text)
        
        if not match:
            await update.message.reply_text(
                "⚠️ Невірний формат!\n\n"
                "Використовуйте: ДД.ММ.РРРР - ДД.ММ.РРРР\n"
                "Наприклад: 01.12.2024 - 31.12.2024"
            )
            return
        
        date_from = match.group(1)
        date_to = match.group(2)
        
        from datetime import datetime
        try:
            datetime.strptime(date_from, "%d.%m.%Y")
            datetime.strptime(date_to, "%d.%m.%Y")
        except ValueError:
            await update.message.reply_text("⚠️ Невірна дата! Перевірте формат.")
            return
        
        instructor_data = get_instructor_by_telegram_id(user_id)
        if not instructor_data:
            await update.message.reply_text("❌ Помилка.")
            return
        
        instructor_id = instructor_data[0]
        period_text = f"{date_from} - {date_to}"
        
        await show_instructor_stats(update, context, instructor_id, date_from, date_to, period_text)
        
    except Exception as e:
        logger.error(f"Error in handle_stats_custom_period: {e}", exc_info=True)
        await update.message.reply_text("❌ Помилка обробки періоду.")

async def show_instructor_stats(update: Update, context: ContextTypes.DEFAULT_TYPE, instructor_id, date_from, date_to, period_text):
    """Показати статистику інструктора"""
    try:
        stats = get_instructor_stats_period(instructor_id, date_from, date_to)
        
        if not stats:
            await update.message.reply_text("❌ Помилка отримання статистики.")
            return
        
        text = f"📊 Статистика {period_text}\n\n"
        text += f"📝 Занять проведено: {stats['total_lessons']}\n"
        text += f"⏱ Годин відпрацьовано: {stats['total_hours']}\n"
        text += f"💰 Заробіток: {stats['earnings']:.0f} грн\n"
        text += f"⭐ Середній рейтинг: {stats['avg_rating']}\n"
        text += f"❌ Скасовано: {stats['cancelled']}\n"
        
        await update.message.reply_text(text)
        
        await show_instructor_stats_menu(update, context)
        
    except Exception as e:
        logger.error(f"Error in show_instructor_stats: {e}", exc_info=True)
        await update.message.reply_text("❌ Помилка.")

async def show_cancellation_history(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Історія скасувань"""
    user_id = update.message.from_user.id
    
    try:
        instructor_data = get_instructor_by_telegram_id(user_id)
        if not instructor_data:
            await update.message.reply_text("❌ Помилка.")
            return
        
        instructor_id = instructor_data[0]
        
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT date, time, student_name, cancelled_by, cancelled_at
                FROM lessons
                WHERE instructor_id = ? AND status = 'cancelled'
                ORDER BY cancelled_at DESC
                LIMIT 10
            """, (instructor_id,))
            
            cancellations = cursor.fetchall()
        
        if not cancellations:
            await update.message.reply_text("📋 Немає скасованих занять.")
            return
        
        text = "❌ *Історія скасувань:*\n\n"
        
        for date, time, student_name, cancelled_by, cancelled_at in cancellations:
            text += f"📅 {date} {time}\n"
            text += f"👤 {student_name}\n"
            text += f"🚫 Скасував: {cancelled_by}\n"
            if cancelled_at:
                text += f"🕐 {cancelled_at[:16]}\n"
            text += "\n"
        
        await update.message.reply_text(text, parse_mode="Markdown")
        
    except Exception as e:
        logger.error(f"Error in show_cancellation_history: {e}", exc_info=True)
        await update.message.reply_text("❌ Помилка.")

# ======================= RATING FUNCTIONS =======================
async def rate_student_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Меню оцінювання учнів"""
    user_id = update.message.from_user.id
    
    try:
        instructor_data = get_instructor_by_telegram_id(user_id)
        if not instructor_data:
            await update.message.reply_text("❌ Ви не інструктор.")
            return
        
        instructor_id = instructor_data[0]
        
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, date, time, student_name, rating, feedback
                FROM lessons
                WHERE instructor_id = ? 
                  AND status = 'completed'
                  AND instructor_rating IS NULL
                ORDER BY date DESC, time DESC
                LIMIT 10
            """, (instructor_id,))
            
            lessons = cursor.fetchall()
        
        if not lessons:
            await update.message.reply_text("📋 Немає занять для оцінювання.")
            return
        
        context.user_data["lessons_to_rate"] = lessons
        context.user_data["state"] = "rating_select_lesson"
        
        text = "⭐ *Оберіть заняття для оцінювання:*\n\n"
        keyboard = []
        
        for i, (lesson_id, date, time, student_name, rating, feedback) in enumerate(lessons, 1):
            text += f"{i}. {date} {time} - {student_name}\n"
            
            if rating and rating > 0:
                stars = "⭐" * rating
                text += f"   Учень оцінив: {stars} ({rating}/5)\n"
                if feedback:
                    text += f"   💬 \"{feedback}\"\n"
            
            text += "\n"
            keyboard.append([KeyboardButton(f"{i}")])
        
        keyboard.append([KeyboardButton("🔙 Назад")])
        
        await update.message.reply_text(
            text,
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
            parse_mode="Markdown"
        )
        
    except Exception as e:
        logger.error(f"Error in rate_student_menu: {e}", exc_info=True)
        await update.message.reply_text("❌ Помилка.")


async def handle_rating_flow(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обробка процесу оцінювання"""
    state = context.user_data.get("state")
    text = update.message.text
    
    if text == "🔙 Назад":
        await start(update, context)
        return
    
    if state == "rating_select_lesson":
        try:
            lesson_index = int(text) - 1
            lessons = context.user_data.get("lessons_to_rate", [])
            
            if lesson_index < 0 or lesson_index >= len(lessons):
                await update.message.reply_text("⚠️ Невірний номер. Спробуйте ще раз:")
                return
            
            selected_lesson = lessons[lesson_index]
            context.user_data["rating_lesson_id"] = selected_lesson[0]
            context.user_data["rating_student_name"] = selected_lesson[3]
            context.user_data["state"] = "rating_give_score"
            
            keyboard = [[KeyboardButton(str(i))] for i in range(1, 6)]
            keyboard.append([KeyboardButton("🔙 Назад")])
            
            await update.message.reply_text(
                f"⭐ Оцініть учня *{selected_lesson[3]}*\n\n"
                f"Виберіть оцінку від 1 до 5:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
                parse_mode="Markdown"
            )
            
        except ValueError:
            await update.message.reply_text("⚠️ Введіть номер заняття:")
            return
    
    elif state == "rating_give_score":
        try:
            rating = int(text)
            if rating < 1 or rating > 5:
                await update.message.reply_text("⚠️ Оцінка має бути від 1 до 5:")
                return
            
            context.user_data["rating_score"] = rating
            context.user_data["state"] = "rating_give_feedback"
            
            keyboard = [
                [KeyboardButton("➡️ Пропустити")],
                [KeyboardButton("🔙 Назад")]
            ]
            
            await update.message.reply_text(
                "💬 Додайте коментар (або натисніть 'Пропустити'):",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
            
        except ValueError:
            await update.message.reply_text("⚠️ Введіть число від 1 до 5:")
            return
    
    elif state == "rating_give_feedback":
        feedback = "" if text == "➡️ Пропустити" else text
        
        lesson_id = context.user_data.get("rating_lesson_id")
        rating = context.user_data.get("rating_score")
        student_name = context.user_data.get("rating_student_name")
        
        if add_instructor_rating(lesson_id, rating, feedback):
            await update.message.reply_text(
                f"✅ Оцінку додано!\n\n"
                f"👤 {student_name}\n"
                f"⭐ Оцінка: {rating}/5"
            )
        else:
            await update.message.reply_text("❌ Помилка збереження оцінки.")
        
        context.user_data.clear()
        await start(update, context)

# ======================= EDIT SCHEDULE =======================
async def handle_edit_schedule(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Коригування графіку інструктором"""
    state = context.user_data.get("state")
    text = update.message.text
    
    if text == "🔙 Назад":
        await start(update, context)
        return
    
    if text == "✏️ Коригувати графік":
        user_id = update.message.from_user.id
        instructor_data = get_instructor_by_telegram_id(user_id)
        
        if not instructor_data:
            await update.message.reply_text("❌ Помилка.")
            return
        
        instructor_id = instructor_data[0]
        
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, date, time, student_name
                FROM lessons
                WHERE instructor_id = ? AND status = 'active'
                ORDER BY date, time
                LIMIT 10
            """, (instructor_id,))
            
            lessons = cursor.fetchall()
        
        if not lessons:
            await update.message.reply_text("📋 Немає занять для коригування.")
            return
        
        context.user_data["lessons_to_edit"] = lessons
        context.user_data["state"] = "edit_schedule_select"
        
        text = "✏️ *Оберіть заняття для зміни:*\n\n"
        keyboard = []
        
        for i, (lesson_id, date, time, student_name) in enumerate(lessons, 1):
            text += f"{i}. {date} {time} - {student_name}\n"
            keyboard.append([KeyboardButton(f"{i}")])
        
        keyboard.append([KeyboardButton("🔙 Назад")])
        
        await update.message.reply_text(
            text,
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
            parse_mode="Markdown"
        )
        return
    
    if state == "edit_schedule_select":
        try:
            lesson_index = int(text) - 1
            lessons = context.user_data.get("lessons_to_edit", [])
            
            if lesson_index < 0 or lesson_index >= len(lessons):
                await update.message.reply_text("⚠️ Невірний номер:")
                return
            
            selected_lesson = lessons[lesson_index]
            context.user_data["edit_lesson_id"] = selected_lesson[0]
            context.user_data["state"] = "edit_schedule_date"
            
            await update.message.reply_text(
                f"📅 Введіть нову дату у форматі *ДД.ММ.РРРР*\n"
                f"Поточна: {selected_lesson[1]}",
                parse_mode="Markdown"
            )
            
        except ValueError:
            await update.message.reply_text("⚠️ Введіть номер:")
            return
    
    elif state == "edit_schedule_date":
        if not validate_date_format(text):
            await update.message.reply_text("⚠️ Невірний формат. Використовуйте ДД.ММ.РРРР:")
            return
        
        context.user_data["edit_new_date"] = text
        context.user_data["state"] = "edit_schedule_time"
        
        await update.message.reply_text("🕐 Введіть новий час у форматі *ГГ:ХХ*", parse_mode="Markdown")
    
    elif state == "edit_schedule_time":
        if not re.match(r'^\d{1,2}:\d{2}$', text):
            await update.message.reply_text("⚠️ Невірний формат. Використовуйте ГГ:ХХ:")
            return
        
        lesson_id = context.user_data.get("edit_lesson_id")
        new_date = context.user_data.get("edit_new_date")
        new_time = text
        
        if update_lesson(lesson_id, date=new_date, time=new_time):
            await update.message.reply_text(
                f"✅ Графік оновлено!\n\n"
                f"📅 Нова дата: {new_date}\n"
                f"🕐 Новий час: {new_time}"
            )
        else:
            await update.message.reply_text("❌ Помилка оновлення.")
        
        context.user_data.clear()
        await start(update, context)

# ======================= SCHEDULE MANAGEMENT =======================
async def manage_schedule(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Управління графіком"""
    keyboard = [
        [KeyboardButton("🔴 Заблокувати час")],
        [KeyboardButton("🟢 Розблокувати час")],
        [KeyboardButton("📋 Мої блокування")],
        [KeyboardButton("🔙 Назад")]
    ]
    
    context.user_data["state"] = "schedule_menu"
    
    await update.message.reply_text(
        "⚙️ *Управління графіком*\n\nОберіть дію:",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        parse_mode="Markdown"
    )

async def handle_schedule_management(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обробка управління графіком"""
    text = update.message.text
    state = context.user_data.get("state")

    # Обробка розблокування через кнопки з номерами
    if state == "waiting_unblock":
        if text == "🔙 Назад":
            await manage_schedule(update, context)
            return
        
        if text.startswith("❌ "):
            num = text.replace("❌ ", "").strip()
            user_id_s = update.message.from_user.id
            
            # Спочатку з user_data
            blocks_map = context.user_data.get("blocks_map", {})
            
            # Якщо немає — читаємо з БД (інший інстанс Render)
            if not blocks_map:
                import json as _json
                with get_db() as conn:
                    cursor = conn.cursor()
                    try:
                        cursor.execute("""
                            SELECT data FROM user_sessions
                            WHERE user_id = ? AND state = 'waiting_unblock'
                        """, (user_id_s,))
                        row = cursor.fetchone()
                        if row:
                            blocks_map = _json.loads(row[0])
                    except Exception as e:
                        logger.warning(f"session read error: {e}")
            
            block_id = blocks_map.get(num)
            if block_id:
                from database import remove_schedule_block
                if remove_schedule_block(block_id):
                    await update.message.reply_text("✅ Блокування видалено!")
                else:
                    await update.message.reply_text("❌ Помилка видалення.")
                # Очищаємо сесію
                with get_db() as conn:
                    cursor = conn.cursor()
                    try:
                        cursor.execute("DELETE FROM user_sessions WHERE user_id = ?", (user_id_s,))
                        conn.commit()
                    except:
                        pass
                context.user_data.pop("blocks_map", None)
                await manage_schedule(update, context)
            else:
                await update.message.reply_text("⚠️ Невірний номер. Оберіть зі списку.")
            return
        
        return

    if text == "🔙 Назад":
        if state == "schedule_menu":
            await start(update, context)
        else:
            await manage_schedule(update, context)
        return
    
    if text == "🔴 Заблокувати час":
        context.user_data["state"] = "block_choose_date"
        
        dates = get_next_dates(30)
        
        keyboard = []
        for i in range(0, len(dates), 2):
            row = [KeyboardButton(dates[i])]
            if i + 1 < len(dates):
                row.append(KeyboardButton(dates[i + 1]))
            keyboard.append(row)
        
        keyboard.append([KeyboardButton("🔙 Назад")])
        
        await update.message.reply_text(
            "📅 Оберіть дату для блокування (доступно на місяць вперед):",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return
    
    elif text == "🟢 Розблокувати час":
        await show_blocks_to_unblock(update, context)
        context.user_data["state"] = "waiting_unblock"
        return
    
    elif text == "📋 Мої блокування":
        await show_all_blocks(update, context)
        return
    
    if state == "block_choose_date":
        date_parts = text.split()
        
        if len(date_parts) >= 3 and date_parts[0] in ["🟦", "🟥"]:
            date_str = date_parts[2]
        elif len(date_parts) >= 2:
            date_str = date_parts[1]
        else:
            date_str = text
        
        if not validate_date_format(date_str):
            await update.message.reply_text("⚠️ Невірний формат. Оберіть дату з меню.")
            return
        
        context.user_data["block_date"] = date_str
        context.user_data["state"] = "block_choose_time_start"
        
        date_obj = datetime.strptime(date_str, "%d.%m.%Y")
        now = datetime.now(TZ)
        is_today = date_obj.date() == now.date()
        current_hour = now.hour
        
        keyboard = []
        for hour in range(WORK_HOURS_START, WORK_HOURS_END + 1):
            if is_today and hour <= current_hour:
                continue
            keyboard.append([KeyboardButton(f"{hour:02d}:00")])
        
        keyboard.append([KeyboardButton("🔙 Назад")])
        
        await update.message.reply_text(
            "🕐 Оберіть час початку блокування:",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return
    
    elif state == "block_choose_time_start":
        if not re.match(r'^\d{1,2}:\d{2}$', text):
            await update.message.reply_text("⚠️ Невірний формат. Оберіть час з меню.")
            return
        
        context.user_data["block_time_start"] = text
        context.user_data["state"] = "block_choose_time_end"
        
        start_hour = int(text.split(':')[0])
        keyboard = []
        for hour in range(start_hour + 1, WORK_HOURS_END + 2):
            keyboard.append([KeyboardButton(f"{hour:02d}:00")])
        
        keyboard.append([KeyboardButton("🔙 Назад")])
        
        await update.message.reply_text(
            "🕐 Оберіть час кінця блокування:",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return
    
    elif state == "block_choose_time_end":
        if not re.match(r'^\d{1,2}:\d{2}$', text):
            await update.message.reply_text("⚠️ Невірний формат. Оберіть час з меню.")
            return
        
        context.user_data["block_time_end"] = text
        context.user_data["state"] = "block_choose_reason"
        
        keyboard = [
            [KeyboardButton("➡️ Пропустити")],
            [KeyboardButton("🔙 Назад")]
        ]
        
        await update.message.reply_text(
            "💬 Введіть причину блокування (або пропустіть):",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return
    
    elif state == "block_choose_reason":
        reason = "" if text == "➡️ Пропустити" else text
        
        user_id = update.message.from_user.id
        instructor_data = get_instructor_by_telegram_id(user_id)
        
        if not instructor_data:
            await update.message.reply_text("❌ Помилка.")
            return
        
        instructor_id = instructor_data[0]
        block_date = context.user_data["block_date"]
        time_start = context.user_data["block_time_start"]
        time_end = context.user_data["block_time_end"]
        
        date_formatted = block_date
        
        def time_to_minutes(time_str):
            h, m = map(int, time_str.split(':'))
            return h * 60 + m
        
        block_start_min = time_to_minutes(time_start)
        block_end_min = time_to_minutes(time_end)
        
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT student_name, student_phone, time, duration, student_tariff
                FROM lessons
                WHERE instructor_id = ? AND date = ? AND status = 'active'
            """, (instructor_id, date_formatted))
            
            lessons = cursor.fetchall()
        
        conflicting_lessons = []
        for student_name, student_phone, lesson_time, duration, tariff in lessons:
            if ':' not in lesson_time:
                continue
            
            lesson_start_min = time_to_minutes(lesson_time)
            
            if "1.5" in duration:
                lesson_duration = 90
            elif "2" in duration:
                lesson_duration = 120
            else:
                lesson_duration = 60
            
            lesson_end_min = lesson_start_min + lesson_duration
            
            if not (block_end_min <= lesson_start_min or block_start_min >= lesson_end_min):
                conflicting_lessons.append({
                    'name': student_name,
                    'phone': student_phone or "немає",
                    'time': lesson_time,
                    'duration': duration,
                    'tariff': tariff or 0
                })
        
        if conflicting_lessons:
            message = f"❌ Не можна заблокувати!\n\n"
            
            for lesson in conflicting_lessons:
                start_h, start_m = map(int, lesson['time'].split(':'))
                if "1.5" in lesson['duration']:
                    end_h, end_m = start_h + 1, start_m + 30
                elif "2" in lesson['duration']:
                    end_h, end_m = start_h + 2, start_m
                else:
                    end_h, end_m = start_h + 1, start_m
                
                message += f"📅 {block_date}, 🕐 {lesson['time']}-{end_h:02d}:{end_m:02d}\n"
                message += f"👤 {lesson['name']} ({lesson['phone']})\n"
                message += f"💵 {lesson['tariff']} грн, {lesson['duration']}\n\n"
            
            message += "Зв'яжіться з учнем для перенесення."
            
            await update.message.reply_text(message)
            context.user_data.clear()
            await manage_schedule(update, context)
            return
        
        date_for_block = datetime.strptime(block_date, "%d.%m.%Y").strftime("%Y-%m-%d")

        from database import add_schedule_block
        
        if add_schedule_block(instructor_id, date_for_block, time_start, time_end, "blocked", reason):
            await update.message.reply_text(
                f"✅ Час заблоковано!\n\n"
                f"📅 {block_date}\n"
                f"🕐 {time_start} - {time_end}"
            )
        else:
            await update.message.reply_text("❌ Помилка блокування.")
        
        context.user_data.clear()
        await manage_schedule(update, context)

async def show_blocks_to_unblock(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показати блокування для видалення через ReplyKeyboard з номерами"""
    user_id = update.message.from_user.id
    
    try:
        instructor_data = get_instructor_by_telegram_id(user_id)
        if not instructor_data:
            await update.message.reply_text("❌ Помилка.")
            return
        
        instructor_id = instructor_data[0]
        
        today_date = datetime.now(TZ).date()
        today_str = today_date.strftime('%Y-%m-%d')
        
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, date, time_start, time_end, reason
                FROM schedule_blocks
                WHERE instructor_id = ?
                AND date >= ?
                ORDER BY date, time_start
                LIMIT 20
            """, (instructor_id, today_str))
            
            future_blocks = cursor.fetchall()
        
        if not future_blocks:
            await update.message.reply_text("📋 Немає майбутніх блокувань.")
            await manage_schedule(update, context)
            return
        
        # Зберігаємо map номер->block_id в user_data ТА в БД
        blocks_map = {}
        for i, (block_id, date, time_start, time_end, reason) in enumerate(future_blocks, 1):
            blocks_map[str(i)] = block_id
        
        context.user_data["blocks_map"] = blocks_map
        
        # Зберігаємо в БД для надійності між інстансами
        import json
        with get_db() as conn:
            cursor = conn.cursor()
            try:
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS user_sessions (
                        user_id INTEGER PRIMARY KEY,
                        state TEXT,
                        data TEXT,
                        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                cursor.execute("""
                    INSERT OR REPLACE INTO user_sessions (user_id, state, data, updated_at)
                    VALUES (?, ?, ?, CURRENT_TIMESTAMP)
                """, (user_id, "waiting_unblock", json.dumps(blocks_map)))
                conn.commit()
            except Exception as e:
                logger.warning(f"Could not save session: {e}")
        
        text = "🟢 *Оберіть номер блокування для видалення:*\n\n"
        keyboard = []
        
        for i, (block_id, date, time_start, time_end, reason) in enumerate(future_blocks, 1):
            text += f"{i}. 📅 {date} 🕐 {time_start}-{time_end}"
            if reason:
                text += f" — {reason}"
            text += "\n"
            keyboard.append([KeyboardButton(f"❌ {i}")])
        
        keyboard.append([KeyboardButton("🔙 Назад")])
        
        await update.message.reply_text(
            text,
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
            parse_mode="Markdown"
        )
        
    except Exception as e:
        logger.error(f"Error in show_blocks_to_unblock: {e}", exc_info=True)
        await update.message.reply_text("❌ Помилка.")

async def show_all_blocks(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показати всі блокування"""
    user_id = update.message.from_user.id
    
    try:
        instructor_data = get_instructor_by_telegram_id(user_id)
        if not instructor_data:
            await update.message.reply_text("❌ Помилка.")
            return
        
        instructor_id = instructor_data[0]
        today_date = datetime.now(TZ).date()
        
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT date, time_start, time_end, reason
                FROM schedule_blocks
                WHERE instructor_id = ?
                ORDER BY date DESC, time_start DESC
            """, (instructor_id,))
            
            all_blocks = cursor.fetchall()
        
        if not all_blocks:
            await update.message.reply_text("📋 У вас немає заблокованих годин.")
            return
        
        future_blocks = []
        past_blocks = []
        
        for block in all_blocks:
            date_str, time_start, time_end, reason = block
            try:
                block_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                if block_date >= today_date:
                    future_blocks.append(block)
                else:
                    past_blocks.append(block)
            except ValueError:
                continue
        
        future_blocks.sort(key=lambda x: datetime.strptime(x[0], '%Y-%m-%d'))
        
        if not future_blocks and not past_blocks:
            await update.message.reply_text("📋 У вас немає заблокованих годин.")
            return
        
        text = ""
        
        if future_blocks:
            text += "🟢 *Майбутні блокування:*\n"
            current_date = None
            for date, time_start, time_end, reason in future_blocks:
                if date != current_date:
                    text += f"\n📅 *{date}*\n"
                    current_date = date
                text += f"🕐 {time_start} - {time_end}"
                if reason:
                    text += f" | {reason}"
                text += "\n"
        
        if past_blocks:
            text += "\n🔴 *Минулі блокування:*\n"
            current_date = None
            for date, time_start, time_end, reason in past_blocks[:30]:
                if date != current_date:
                    text += f"\n📅 {date}\n"
                    current_date = date
                text += f"🕐 {time_start} - {time_end}"
                if reason:
                    text += f" | {reason}"
                text += "\n"
        
        await update.message.reply_text(text, parse_mode="Markdown")
        
    except Exception as e:
        logger.error(f"Error in show_all_blocks: {e}", exc_info=True)
        await update.message.reply_text("❌ Помилка.")

# ======================= ADMIN FUNCTIONS =======================
async def show_admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Панель адміністратора"""
    keyboard = [
        [KeyboardButton("📊 Звіт по інструкторах")],
        [KeyboardButton("👤 Звіт по інструктору")],
        [KeyboardButton("👥 Список інструкторів")],
        [KeyboardButton("✏️ Управління записами")],
        [KeyboardButton("➕ Додати учня")],
        [KeyboardButton("📥 Експорт в Excel")],
        [KeyboardButton("🔙 Назад")]
    ]
    
    await update.message.reply_text(
        "🔐 Панель адміністратора\n\nОберіть дію:",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    )
    
    context.user_data["state"] = "admin_panel"

async def handle_admin_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обробка звітів адміна"""
    text = update.message.text
    
    if text == "🔙 Назад":
        context.user_data.clear()
        await start(update, context)
        return
    
    if text == "➕ Додати учня":
        await admin_add_student_start(update, context)
        return
    
    if text == "✏️ Управління записами":
        keyboard = [
            [KeyboardButton("❌ Скасувати запис учня")],
            [KeyboardButton("🔙 Назад")]
        ]
        context.user_data["state"] = "admin_manage_bookings"
        await update.message.reply_text(
            "✏️ Управління записами\n\nОберіть дію:",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return
    
    if text == "👤 Звіт по інструктору":
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id, name FROM instructors ORDER BY name")
            instructors = cursor.fetchall()
        
        keyboard = []
        # ✅ Кнопка "Всі інструктори" на першому місці
        keyboard.append([KeyboardButton("👥 Всі інструктори")])
        for inst_id, inst_name in instructors:
            keyboard.append([KeyboardButton(f"👤 {inst_name}")])
        keyboard.append([KeyboardButton("🔙 Назад")])
        
        context.user_data["state"] = "admin_select_instructor_report"
        context.user_data["instructor_list"] = {inst_name: inst_id for inst_id, inst_name in instructors}
        
        await update.message.reply_text(
            "👤 Оберіть інструктора:",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return

    if text == "📊 Звіт по інструкторах":
        keyboard = [
            [KeyboardButton("📊 За тиждень")],
            [KeyboardButton("📊 За місяць")],
            [KeyboardButton("📊 Свій період")],
            [KeyboardButton("🔙 Назад")]
        ]
        
        context.user_data["state"] = "admin_report_period"
        
        await update.message.reply_text(
            "📊 Оберіть період для звіту:",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return
    
    if text == "👥 Список інструкторів":
        instructors = get_all_instructors()
        
        text = "👥 Список інструкторів:\n\n"
        for i, (inst_id, name, transmission, telegram_id) in enumerate(instructors, 1):
            text += f"{i}. {name} ({transmission})\n"
            text += f"   ID: {telegram_id}\n\n"
        
        await update.message.reply_text(text)
        await show_admin_panel(update, context)
        return
    
    today = datetime.now().date()
    
    if text == "📊 За тиждень":
        date_from = (today - timedelta(days=7)).strftime("%Y-%m-%d")
        date_to = today.strftime("%Y-%m-%d")
        period_text = "за тиждень"
    elif text == "📊 За місяць":
        date_from = (today - timedelta(days=30)).strftime("%Y-%m-%d")
        date_to = today.strftime("%Y-%m-%d")
        period_text = "за місяць"
    elif text == "📊 Свій період":
        await update.message.reply_text(
            "📅 Введіть період у форматі:\n"
            "ДД.ММ.РРРР - ДД.ММ.РРРР\n\n"
            "Наприклад: 01.11.2024 - 30.11.2024"
        )
        context.user_data["state"] = "admin_custom_period"
        return
    else:
        return
    
    await generate_admin_report(update, context, date_from, date_to, period_text)

async def handle_instructor_custom_period(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обробка свого періоду для звіту по інструктору"""
    text = update.message.text
    
    if text == "🔙 Назад":
        context.user_data["state"] = "admin_panel"
        await show_admin_panel(update, context)
        return
    
    try:
        parts = text.split(" - ")
        date_from = datetime.strptime(parts[0].strip(), "%d.%m.%Y").strftime("%Y-%m-%d")
        date_to = datetime.strptime(parts[1].strip(), "%d.%m.%Y").strftime("%Y-%m-%d")
        period_text = f"за період {parts[0].strip()} - {parts[1].strip()}"
        
        instructor_id = context.user_data.get("selected_instructor_id")
        instructor_name = context.user_data.get("selected_instructor_name")

        # ✅ Розгалуження: один чи всі
        if instructor_id == "all":
            await generate_all_instructors_report(update, context, date_from, date_to, period_text)
        else:
            await generate_instructor_report(update, context, instructor_id, instructor_name, date_from, date_to, period_text)
    except Exception as e:
        await update.message.reply_text("❌ Невірний формат. Введіть: ДД.ММ.РРРР - ДД.ММ.РРРР")

async def handle_instructor_report_select(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обробка вибору інструктора для звіту"""
    text = update.message.text
    
    if text == "🔙 Назад":
        context.user_data["state"] = "admin_panel"
        await show_admin_panel(update, context)
        return
    
    # ✅ Обробка "Всі інструктори"
    if text == "👥 Всі інструктори":
        context.user_data["selected_instructor_id"] = "all"
        context.user_data["selected_instructor_name"] = "Всі інструктори"
        context.user_data["state"] = "admin_instructor_report_period"
        
        keyboard = [
            [KeyboardButton("📊 За тиждень")],
            [KeyboardButton("📊 За місяць")],
            [KeyboardButton("📊 Свій період")],
            [KeyboardButton("🔙 Назад")]
        ]
        await update.message.reply_text(
            "👥 Всі інструктори\nОберіть період:",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return

    # Витягуємо ім'я з кнопки (прибираємо emoji "👤 ")
    instructor_name = text.replace("👤 ", "").strip()
    instructor_list = context.user_data.get("instructor_list", {})
    
    if instructor_name not in instructor_list:
        await update.message.reply_text("❌ Інструктора не знайдено.")
        return
    
    context.user_data["selected_instructor_id"] = instructor_list[instructor_name]
    context.user_data["selected_instructor_name"] = instructor_name
    context.user_data["state"] = "admin_instructor_report_period"
    
    keyboard = [
        [KeyboardButton("📊 За тиждень")],
        [KeyboardButton("📊 За місяць")],
        [KeyboardButton("📊 Свій період")],
        [KeyboardButton("🔙 Назад")]
    ]
    await update.message.reply_text(
        f"👤 {instructor_name}\nОберіть період:",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    )

async def handle_instructor_report_period(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обробка вибору періоду для звіту по інструктору"""
    text = update.message.text
    
    if text == "🔙 Назад":
        context.user_data["state"] = "admin_panel"
        await show_admin_panel(update, context)
        return
    
    today = datetime.now().date()
    
    if text == "📊 За тиждень":
        date_from = (today - timedelta(days=7)).strftime("%Y-%m-%d")
        date_to = today.strftime("%Y-%m-%d")
        period_text = "за тиждень"
    elif text == "📊 За місяць":
        date_from = (today - timedelta(days=30)).strftime("%Y-%m-%d")
        date_to = today.strftime("%Y-%m-%d")
        period_text = "за місяць"
    elif text == "📊 Свій період":
        await update.message.reply_text(
            "📅 Введіть період у форматі:\nДД.ММ.РРРР - ДД.ММ.РРРР\n\nНаприклад: 01.11.2024 - 30.11.2024"
        )
        context.user_data["state"] = "admin_instructor_custom_period"
        return
    else:
        return
    
    instructor_id = context.user_data.get("selected_instructor_id")
    instructor_name = context.user_data.get("selected_instructor_name")

    # ✅ Розгалуження: один чи всі
    if instructor_id == "all":
        await generate_all_instructors_report(update, context, date_from, date_to, period_text)
    else:
        await generate_instructor_report(update, context, instructor_id, instructor_name, date_from, date_to, period_text)

# ======================= НОВИЙ ЗВІТ: ВСІ ІНСТРУКТОРИ =======================
async def generate_all_instructors_report(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    date_from: str,
    date_to: str,
    period_text: str
):
    """Детальний звіт по ВСІХ інструкторах за період (з розбивкою по днях)"""
    try:
        period_from = datetime.strptime(date_from, "%Y-%m-%d").strftime("%d.%m.%Y")
        period_to   = datetime.strptime(date_to,   "%Y-%m-%d").strftime("%d.%m.%Y")

        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id, name FROM instructors WHERE is_active = 1 ORDER BY name")
            instructors = cursor.fetchall()

        if not instructors:
            await update.message.reply_text("📋 Інструкторів не знайдено.")
            await show_admin_panel(update, context)
            return

        # Заголовок
        header = (
            f"📊 *Звіт по всіх інструкторах*\n"
            f"📅 {period_from} – {period_to}\n"
        )
        await update.message.reply_text(header, parse_mode="Markdown")

        CLEAN_RATE = 420   # чистий дохід за годину
        AMORT_RATE = 70    # амортизація за годину

        grand_lessons  = 0
        grand_hours    = 0.0
        grand_zagalno  = 0.0
        grand_clean    = 0.0
        grand_amort    = 0.0

        for inst_id, inst_name in instructors:
            data = get_instructor_report(inst_id, date_from, date_to)
            if not data or data["total_lessons"] == 0:
                continue  # Пропускаємо інструкторів без занять за період

            # Рахуємо загальну суму по тарифах учнів з БД
            from database import _date_in_range
            with get_db() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT date, duration, student_tariff
                    FROM lessons
                    WHERE instructor_id = ?
                    AND status IN ('active', 'completed')
                """, (inst_id,))
                cursor2_rows = cursor.fetchall()

            inst_zagalno = 0.0
            for row_date, row_dur, row_tariff in cursor2_rows:
                if not _date_in_range(row_date, date_from, date_to):
                    continue
                h = 1.5 if row_dur and '1.5' in str(row_dur) else (2.0 if row_dur and '2' in str(row_dur) else 1.0)
                tariff = row_tariff or 490
                inst_zagalno += h * tariff

            inst_hours  = data["total_hours"]
            inst_clean  = inst_hours * CLEAN_RATE
            inst_amort  = inst_zagalno - inst_clean

            grand_lessons += data["total_lessons"]
            grand_hours   += inst_hours
            grand_zagalno += inst_zagalno
            grand_clean   += inst_clean
            grand_amort   += inst_amort

            # Блок одного інструктора
            lines = [f"━━━ 👨‍🏫 *{inst_name}* ━━━"]

            if data["details"]:
                current_date = None
                for det_date, det_time, det_hours, student_name, status, rating in data["details"]:
                    if det_date != current_date:
                        lines.append(f"\n📆 *{det_date}*")
                        current_date = det_date

                    status_icon = (
                        "✅" if status == "completed" else
                        "❌" if status == "cancelled" else
                        "🔵"
                    )
                    line = f"{status_icon} {det_time} ({det_hours}г) – {student_name}"
                    if rating:
                        line += f" ⭐{rating}"
                    lines.append(line)

            # Підсумок по інструктору
            avg = data["avg_rating"]
            rating_str = f"{avg:.1f}" if avg else "–"
            lines.append(
                f"\n📝 {data['total_lessons']} | ⏱ {inst_hours}г | ⭐ {rating_str}\n"
                f"💰 Загально: {inst_zagalno:.0f} грн\n"
                f"💵 Чистими: {inst_clean:.0f} грн\n"
                f"🔧 Амортизація: {inst_amort:.0f} грн"
            )

            # Розбиваємо на шматки по 3500 символів (ліміт Telegram ~4096)
            block = "\n".join(lines)
            chunks = []
            while len(block) > 3500:
                split_at = block.rfind("\n", 0, 3500)
                if split_at == -1:
                    split_at = 3500
                chunks.append(block[:split_at])
                block = block[split_at:]
            chunks.append(block)

            for chunk in chunks:
                if chunk.strip():
                    await update.message.reply_text(chunk.strip(), parse_mode="Markdown")

        # Загальний підсумок
        if grand_lessons == 0:
            summary = f"📋 За цей період занять не знайдено."
        else:
            summary = (
                f"──────────────────\n"
                f"📊 *РАЗОМ {period_text}:*\n"
                f"📝 {grand_lessons} занять  |  ⏱ {grand_hours:.1f} год\n"
                f"💰 Загально: {grand_zagalno:.0f} грн\n"
                f"💵 Чистими: {grand_clean:.0f} грн\n"
                f"🔧 Амортизація: {grand_amort:.0f} грн"
            )
        await update.message.reply_text(summary, parse_mode="Markdown")
        await show_admin_panel(update, context)

    except Exception as e:
        logger.error(f"Error in generate_all_instructors_report: {e}", exc_info=True)
        await update.message.reply_text("❌ Помилка генерації звіту.")
        await show_admin_panel(update, context)

# ======================= ОДИНОЧНИЙ ЗВІТ ПО ІНСТРУКТОРУ =======================
async def generate_instructor_report(update: Update, context: ContextTypes.DEFAULT_TYPE, instructor_id, instructor_name, date_from, date_to, period_text):
    """Генерація детального звіту по одному інструктору"""
    try:
        data = get_instructor_report(instructor_id, date_from, date_to)
        
        if not data:
            await update.message.reply_text("❌ Помилка отримання даних.")
            return
        
        period_from = datetime.strptime(date_from, "%Y-%m-%d").strftime("%d.%m.%Y")
        period_to = datetime.strptime(date_to, "%Y-%m-%d").strftime("%d.%m.%Y")
        
        text = f"👤 *{instructor_name}*\n"
        text += f"📅 Період: {period_from} - {period_to}\n\n"
        text += f"📝 Занять: {data['total_lessons']}\n"
        text += f"⏱ Годин: {data['total_hours']}\n"
        text += f"💰 Заробіток: {data['earnings']:.0f} грн\n"
        text += f"⭐ Рейтинг: {data['avg_rating']:.1f}\n"
        text += f"❌ Скасовано: {data['cancelled']}\n"
        
        if data['details']:
            text += "\n📋 *Деталі занять:*\n"
            current_date = None
            for date, time, hours, student_name, status, rating in data['details']:
                if date != current_date:
                    text += f"\n📆 *{date}*\n"
                    current_date = date
                status_icon = "✅" if status == "completed" else ("❌" if status == "cancelled" else "🔵")
                text += f"{status_icon} {time} ({hours}г) - {student_name}"
                if rating:
                    text += f" ⭐{rating}"
                text += "\n"
        
        await update.message.reply_text(text, parse_mode="Markdown")
        await show_admin_panel(update, context)
        
    except Exception as e:
        logger.error(f"Error in generate_instructor_report: {e}", exc_info=True)
        await update.message.reply_text("❌ Помилка генерації звіту.")

async def generate_admin_report(update: Update, context: ContextTypes.DEFAULT_TYPE, date_from, date_to, period_text):
    """Генерація звіту для адміна"""
    try:
        report_data = get_admin_report_by_instructors(date_from, date_to)
        
        if not report_data:
            await update.message.reply_text("📋 Немає даних за цей період.")
            return
        
        text = f"📊 Звіт по інструкторах {period_text}\n\n"
        text += f"📅 Період: {date_from} - {date_to}\n\n"
        
        total_lessons = 0
        total_hours = 0
        total_earnings = 0
        
        for name, lessons, hours, avg_rating, cancelled in report_data:
            if lessons > 0:
                hours = hours or 0
                earnings = hours * 420
                
                text += f"👨‍🏫 {name}\n"
                text += f"   📝 Занять: {lessons}\n"
                text += f"   ⏱ Годин: {hours:.1f}\n"
                text += f"   💰 Заробіток: {earnings:.0f} грн\n"
                text += f"   ⭐ Рейтинг: {avg_rating:.1f if avg_rating else '0.0'}\n"
                text += f"   ❌ Скасовано: {cancelled}\n\n"
                
                total_lessons += lessons
                total_hours += hours
                total_earnings += earnings
        
        text += f"\n📊 ЗАГАЛОМ:\n"
        text += f"📝 Занять: {total_lessons}\n"
        text += f"⏱ Годин: {total_hours:.1f}\n"
        text += f"💰 Заробіток: {total_earnings:.0f} грн\n"
        
        await update.message.reply_text(text)
        await show_admin_panel(update, context)
        
    except Exception as e:
        logger.error(f"Error in generate_admin_report: {e}", exc_info=True)
        await update.message.reply_text("❌ Помилка генерації звіту.")

# ======================= ADMIN ADD STUDENT =======================
async def admin_add_student_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Адмін розпочинає додавання нового учня"""
    if not is_admin(update.message.from_user.id):
        await update.message.reply_text("❌ Доступ заборонено.")
        return
    
    context.user_data["state"] = "admin_add_student_name"
    context.user_data["new_student"] = {}
    
    keyboard = [[KeyboardButton("🔙 Скасувати")]]
    await update.message.reply_text(
        "➕ *Додавання нового учня*\n\n"
        "Крок 1 з 4: Введіть ім'я та прізвище учня:\n\n"
        "_Наприклад: Іваненко Олексій_",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        parse_mode="Markdown"
    )

async def handle_admin_add_student_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "🔙 Скасувати":
        await show_admin_panel(update, context)
        return
    
    if len(text.strip()) < 3:
        await update.message.reply_text("⚠️ Введіть повне ім'я (мінімум 3 символи):")
        return
    
    context.user_data["new_student"]["name"] = text.strip()
    context.user_data["state"] = "admin_add_student_phone"
    
    keyboard = [[KeyboardButton("🔙 Скасувати")]]
    await update.message.reply_text(
        f"✅ Ім'я: *{text.strip()}*\n\n"
        "Крок 2 з 4: Введіть номер телефону учня:\n\n"
        "_Формат: +380501234567 або 0501234567_",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        parse_mode="Markdown"
    )

async def handle_admin_add_student_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "🔙 Скасувати":
        await show_admin_panel(update, context)
        return
    
    if not validate_phone(text):
        await update.message.reply_text(
            "⚠️ Невірний формат номера.\n\n"
            "Введіть у форматі: +380501234567 або 0501234567"
        )
        return
    
    existing = get_student_by_phone(text)
    if existing:
        await update.message.reply_text(
            f"⚠️ Учень з таким номером вже є в системі!\n\n"
            f"👤 {existing[1]}\n"
            f"📱 {existing[2]}\n"
            f"💰 Тариф: {existing[3]} грн/год\n\n"
            f"Введіть інший номер або натисніть «Скасувати»:"
        )
        return
    
    context.user_data["new_student"]["phone"] = text.strip()
    context.user_data["state"] = "admin_add_student_tariff"
    
    keyboard = [
        [KeyboardButton("💰 490 грн/год"), KeyboardButton("💰 590 грн/год")],
        [KeyboardButton("🔙 Скасувати")]
    ]
    name = context.user_data["new_student"]["name"]
    await update.message.reply_text(
        f"✅ Ім'я: *{name}*\n"
        f"✅ Телефон: *{text.strip()}*\n\n"
        "Крок 3 з 4: Оберіть тариф учня:",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        parse_mode="Markdown"
    )

async def handle_admin_add_student_tariff(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "🔙 Скасувати":
        await show_admin_panel(update, context)
        return
    
    tariff_map = {
        "💰 490 грн/год": 490,
        "💰 590 грн/год": 590
    }
    
    if text not in tariff_map:
        await update.message.reply_text("⚠️ Оберіть тариф із меню.")
        return
    
    tariff = tariff_map[text]
    context.user_data["new_student"]["tariff"] = tariff
    context.user_data["state"] = "admin_add_student_tgid"
    
    keyboard = [
        [KeyboardButton("⏭️ Пропустити")],
        [KeyboardButton("🔙 Скасувати")]
    ]
    name = context.user_data["new_student"]["name"]
    phone = context.user_data["new_student"]["phone"]
    await update.message.reply_text(
        f"✅ Ім'я: *{name}*\n"
        f"✅ Телефон: *{phone}*\n"
        f"✅ Тариф: *{tariff} грн/год*\n\n"
        "Крок 4 з 4: Введіть Telegram ID учня\n\n"
        "_(якщо учень вже писав боту — попросіть його переслати будь-яке повідомлення боту, "
        "або знайдіть ID у логах. Якщо не знаєте — пропустіть)_",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        parse_mode="Markdown"
    )

async def handle_admin_add_student_tgid(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "🔙 Скасувати":
        await show_admin_panel(update, context)
        return
    
    student_data = context.user_data.get("new_student", {})
    name = student_data.get("name", "")
    phone = student_data.get("phone", "")
    tariff = student_data.get("tariff", 490)
    
    telegram_id = None
    
    if text != "⏭️ Пропустити":
        if not text.strip().lstrip("-").isdigit():
            await update.message.reply_text(
                "⚠️ Telegram ID — це число.\n\n"
                "Введіть числовий ID або натисніть «Пропустити»:"
            )
            return
        telegram_id = int(text.strip())
        
        existing = get_student_by_telegram_id(telegram_id)
        if existing:
            await update.message.reply_text(
                f"⚠️ Цей Telegram ID вже прив'язаний до учня:\n\n"
                f"👤 {existing[1]}\n"
                f"📱 {existing[2]}\n\n"
                "Введіть інший ID або натисніть «Пропустити»:"
            )
            return
    
    try:
        success = register_student(name, phone, telegram_id, tariff, "admin")
        
        if success:
            summary = (
                f"✅ *Учня успішно додано!*\n\n"
                f"👤 Ім'я: {name}\n"
                f"📱 Телефон: {phone}\n"
                f"💰 Тариф: {tariff} грн/год\n"
            )
            if telegram_id:
                summary += f"🆔 Telegram ID: {telegram_id}\n"
            else:
                summary += "🆔 Telegram ID: не вказано\n"
            
            await update.message.reply_text(summary, parse_mode="Markdown")
            
            if telegram_id:
                try:
                    await update.get_bot().send_message(
                        chat_id=telegram_id,
                        text=(
                            "✅ *Вас зареєстровано в системі автошколи!*\n\n"
                            f"👤 Ім'я: {name}\n"
                            f"💰 Ваш тариф: {tariff} грн/год\n\n"
                            "Натисніть /start щоб розпочати роботу з ботом."
                        ),
                        parse_mode="Markdown"
                    )
                    await update.message.reply_text(
                        "📤 Учню надіслано повідомлення про реєстрацію."
                    )
                except Exception as e:
                    logger.warning(f"Не вдалось надіслати повідомлення учню {telegram_id}: {e}")
                    await update.message.reply_text(
                        "⚠️ Учня додано, але надіслати повідомлення не вдалося.\n"
                        "Можливо, він ще не писав боту — нагадайте учню написати /start."
                    )
        else:
            await update.message.reply_text("❌ Помилка збереження. Спробуйте ще раз.")
            
    except Exception as e:
        logger.error(f"Error in handle_admin_add_student_tgid: {e}", exc_info=True)
        await update.message.reply_text(f"❌ Помилка: {e}")
    
    context.user_data.pop("new_student", None)
    await show_admin_panel(update, context)

# ======================= ADMIN MANAGE BOOKINGS =======================
async def handle_admin_manage_bookings(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обробка меню управління записами"""
    text = update.message.text
    
    if text == "✏️ Управління записами":
        keyboard = [
            [KeyboardButton("❌ Скасувати запис учня")],
            [KeyboardButton("➕ Записати учня вручну")],
            [KeyboardButton("🔙 Назад")]
        ]
        context.user_data["state"] = "admin_manage_bookings"
        await update.message.reply_text(
            "✏️ Управління записами\n\nОберіть дію:",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return
    
    if text == "🔙 Назад":
        await show_admin_panel(update, context)
        return
    
    if text == "➕ Записати учня вручну":
        context.user_data["state"] = "admin_manual_enter_phone"
        context.user_data["admin_booking"] = {}
        
        await update.message.reply_text(
            "📱 *Крок 1/7: Телефон учня*\n\n"
            "Введіть номер телефону:\n"
            "Формат: +380501234567 або 0501234567",
            reply_markup=ReplyKeyboardMarkup([[KeyboardButton("🔙 Назад")]], resize_keyboard=True),
            parse_mode="Markdown"
        )
        return
    
    if text == "❌ Скасувати запис учня":
        today = datetime.now(TZ).date()
        dates_with_lessons = []
        
        for i in range(-7, 31):
            date = today + timedelta(days=i)
            date_str = date.strftime('%d.%m.%Y')
            
            with get_db() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT COUNT(*) FROM lessons 
                    WHERE date = ? AND status = 'active'
                """, (date_str,))
                count = cursor.fetchone()[0]
            
            if count > 0:
                weekday = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Нд"][date.weekday()]
                formatted = f"{weekday} {date.strftime('%d.%m')} ({count} уроків)"
                dates_with_lessons.append((date_str, formatted))
        
        if not dates_with_lessons:
            await update.message.reply_text("📋 Немає активних записів на найближчі дні.")
            await show_admin_panel(update, context)
            return
        
        keyboard = []
        for date_str, formatted in dates_with_lessons[:20]:
            keyboard.append([KeyboardButton(formatted)])
        keyboard.append([KeyboardButton("🔙 Назад")])
        
        context.user_data["state"] = "admin_cancel_select_date"
        context.user_data["dates_map"] = {formatted: date_str for date_str, formatted in dates_with_lessons}
        
        await update.message.reply_text(
            "📅 Оберіть дату для перегляду уроків:",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )

async def handle_admin_cancel_select_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "🔙 Назад":
        keyboard = [
            [KeyboardButton("❌ Скасувати запис учня")],
            [KeyboardButton("➕ Записати учня вручну")],
            [KeyboardButton("🔙 Назад")]
        ]
        context.user_data["state"] = "admin_manage_bookings"
        await update.message.reply_text(
            "✏️ Управління записами\n\nОберіть дію:",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return
    
    dates_map = context.user_data.get("dates_map", {})
    date_str = dates_map.get(text)
    
    if not date_str:
        await update.message.reply_text("❌ Невірна дата. Оберіть зі списку.")
        return
    
    context.user_data["selected_date"] = date_str
    
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT i.name, COUNT(*) as lesson_count
            FROM lessons l
            JOIN instructors i ON l.instructor_id = i.id
            WHERE l.date = ? AND l.status = 'active'
            GROUP BY i.name
            ORDER BY i.name
        """, (date_str,))
        instructors = cursor.fetchall()
    
    if not instructors:
        await update.message.reply_text("📋 Немає активних уроків на цю дату.")
        return
    
    keyboard = []
    total_lessons = sum(count for _, count in instructors)
    
    for instructor_name, lesson_count in instructors:
        keyboard.append([KeyboardButton(f"👨‍🏫 {instructor_name} ({lesson_count})")])
    
    if total_lessons <= 15:
        keyboard.append([KeyboardButton("📋 Всі уроки")])
    
    keyboard.append([KeyboardButton("🔙 Назад")])
    
    context.user_data["state"] = "admin_cancel_select_instructor"
    
    await update.message.reply_text(
        f"📅 {date_str} ({total_lessons} уроків)\n\nОберіть інструктора:",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    )

async def handle_admin_cancel_select_instructor(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    date_str = context.user_data.get("selected_date")
    
    if text == "🔙 Назад":
        context.user_data["state"] = "admin_cancel_select_date"
        await update.message.reply_text("📅 Оберіть іншу дату:")
        return
    
    if text == "📋 Всі уроки":
        instructor_filter = None
    else:
        instructor_name = text.replace("👨‍🏫 ", "").split(" (")[0].strip()
        instructor_filter = instructor_name
    
    with get_db() as conn:
        cursor = conn.cursor()
        if instructor_filter:
            cursor.execute("""
                SELECT l.id, l.time, l.duration, l.student_name, l.student_phone, i.name
                FROM lessons l
                JOIN instructors i ON l.instructor_id = i.id
                WHERE l.date = ? AND l.status = 'active' AND i.name = ?
                ORDER BY l.time
            """, (date_str, instructor_filter))
        else:
            cursor.execute("""
                SELECT l.id, l.time, l.duration, l.student_name, l.student_phone, i.name
                FROM lessons l
                JOIN instructors i ON l.instructor_id = i.id
                WHERE l.date = ? AND l.status = 'active'
                ORDER BY l.time
            """, (date_str,))
        lessons = cursor.fetchall()
    
    if not lessons:
        await update.message.reply_text("📋 Немає уроків для цього інструктора.")
        return
    
    message = f"📅 {date_str}"
    if instructor_filter:
        message += f" - {instructor_filter}"
    message += f" ({len(lessons)} уроків):\n\n"
    
    keyboard = []
    
    for idx, (lesson_id, time, duration, student_name, student_phone, instructor_name) in enumerate(lessons, 1):
        message += f"{idx}️⃣ {time} {student_name[:15]}\n"
        if not instructor_filter:
            message += f"   👨‍🏫 {instructor_name}\n"
        
        keyboard.append([KeyboardButton(f"{idx}️⃣")])
    
    keyboard.append([KeyboardButton("🔙 Назад")])
    
    context.user_data["state"] = "admin_cancel_select_lesson"
    context.user_data["lessons_on_date"] = {str(idx): lesson_id for idx, (lesson_id, *_) in enumerate(lessons, 1)}
    
    await update.message.reply_text(
        message + "\n💡 Оберіть номер уроку:",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    )

async def handle_admin_cancel_select_lesson(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "🔙 Назад":
        context.user_data["state"] = "admin_cancel_select_instructor"
        await update.message.reply_text("👨‍🏫 Оберіть іншого інструктора:")
        return
    
    lesson_num = text.replace("️⃣", "")
    lessons_on_date = context.user_data.get("lessons_on_date", {})
    lesson_id = lessons_on_date.get(lesson_num)
    
    if not lesson_id:
        await update.message.reply_text("❌ Невірний номер. Оберіть зі списку.")
        return
    
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT l.student_name, l.student_phone, l.date, l.time, l.duration, i.name, l.student_telegram_id
            FROM lessons l
            JOIN instructors i ON l.instructor_id = i.id
            WHERE l.id = ?
        """, (lesson_id,))
        lesson = cursor.fetchone()
    
    if not lesson:
        await update.message.reply_text("❌ Урок не знайдено.")
        return
    
    student_name, student_phone, date, time, duration, instructor_name, student_telegram_id = lesson
    
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE lessons
                SET status = 'cancelled',
                    cancelled_by = 'admin',
                    cancelled_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (lesson_id,))
            conn.commit()
        
        if student_telegram_id:
            try:
                await context.bot.send_message(
                    chat_id=student_telegram_id,
                    text=f"😔 Вибачте, ваш урок на {date} о {time} з інструктором {instructor_name} "
                         f"скасовано адміністратором.\n\n"
                         f"Зв'яжіться з нами для перенесення:\n📞 +380671234567"
                )
            except Exception as e:
                logger.error(f"Не вдалось відправити повідомлення учню {student_telegram_id}: {e}")
        
        await update.message.reply_text(
            f"✅ Урок скасовано!\n\n"
            f"👤 Учень: {student_name}\n"
            f"👨‍🏫 Інструктор: {instructor_name}\n"
            f"📅 {date} {time}\n"
            f"⏱ {duration}\n\n"
            f"{'📱 Учня сповіщено' if student_telegram_id else '⚠️ Учень НЕ має Telegram - зателефонуйте: ' + (student_phone or 'номер невідомий')}"
        )
        
        await show_admin_panel(update, context)
        
    except Exception as e:
        logger.error(f"Error cancelling lesson: {e}")
        await update.message.reply_text("❌ Помилка при скасуванні уроку.")

# ======================= ADMIN MANUAL BOOKING =======================
async def handle_admin_manual_enter_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "🔙 Назад":
        keyboard = [
            [KeyboardButton("❌ Скасувати запис учня")],
            [KeyboardButton("➕ Записати учня вручну")],
            [KeyboardButton("🔙 Назад")]
        ]
        context.user_data["state"] = "admin_manage_bookings"
        await update.message.reply_text(
            "✏️ Управління записами\n\nОберіть дію:",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return
    
    phone = text.strip()
    if phone.startswith("0"):
        phone = "+38" + phone
    elif not phone.startswith("+"):
        phone = "+" + phone
    
    if not re.match(r'^\+380\d{9}$', phone):
        await update.message.reply_text(
            "❌ Невірний формат!\n\n"
            "Використайте:\n"
            "+380501234567 або 0501234567"
        )
        return
    
    context.user_data["admin_booking"]["phone"] = phone
    
    student = get_student_by_phone(phone)
    
    if student:
        student_id, name, student_phone, tariff, registered_via, student_tg_id = student
        context.user_data["admin_booking"]["name"] = name
        context.user_data["admin_booking"]["tariff"] = tariff
        context.user_data["admin_booking"]["existing_student"] = True
        context.user_data["admin_booking"]["student_telegram_id"] = student_tg_id
        
        keyboard = [
            [KeyboardButton("✅ Так, це той учень")],
            [KeyboardButton("✏️ Ні, ввести дані вручну")],
            [KeyboardButton("🔙 Назад")]
        ]
        
        await update.message.reply_text(
            f"✅ *Знайдено учня:*\n\n"
            f"👤 Ім'я: {name}\n"
            f"📱 Телефон: {phone}\n"
            f"💰 Тариф: {tariff} грн/год\n\n"
            f"Підтвердити?",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
            parse_mode="Markdown"
        )
        context.user_data["state"] = "admin_manual_confirm_student"
    else:
        context.user_data["admin_booking"]["existing_student"] = False
        context.user_data["state"] = "admin_manual_enter_name"
        
        await update.message.reply_text(
            f"❌ *Учня не знайдено*\n\n"
            f"📱 Телефон: {phone}\n\n"
            f"📝 *Крок 2/7: Ім'я учня*\n"
            f"Введіть ім'я та прізвище:",
            reply_markup=ReplyKeyboardMarkup([[KeyboardButton("🔙 Назад")]], resize_keyboard=True),
            parse_mode="Markdown"
        )

async def handle_admin_manual_confirm_student(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "🔙 Назад":
        context.user_data["state"] = "admin_manual_enter_phone"
        await update.message.reply_text(
            "📱 Введіть номер телефону:",
            reply_markup=ReplyKeyboardMarkup([[KeyboardButton("🔙 Назад")]], resize_keyboard=True)
        )
        return
    
    if text == "✅ Так, це той учень":
        keyboard = [
            [KeyboardButton("🚗 Автомат"), KeyboardButton("🚙 Механіка")],
            [KeyboardButton("🔙 Назад")]
        ]
        context.user_data["state"] = "admin_manual_select_transmission"
        await update.message.reply_text(
            f"🚗 *Крок 3/7: Тип коробки*\n\nОберіть тип коробки передач:",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
            parse_mode="Markdown"
        )
    elif text == "✏️ Ні, ввести дані вручну":
        context.user_data["admin_booking"]["existing_student"] = False
        context.user_data["state"] = "admin_manual_enter_name"
        await update.message.reply_text(
            "📝 *Крок 2/7: Ім'я*\nВведіть ім'я та прізвище:",
            reply_markup=ReplyKeyboardMarkup([[KeyboardButton("🔙 Назад")]], resize_keyboard=True),
            parse_mode="Markdown"
        )

async def handle_admin_manual_enter_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "🔙 Назад":
        context.user_data["state"] = "admin_manual_enter_phone"
        await update.message.reply_text(
            "📱 Введіть номер телефону:",
            reply_markup=ReplyKeyboardMarkup([[KeyboardButton("🔙 Назад")]], resize_keyboard=True)
        )
        return
    
    context.user_data["admin_booking"]["name"] = text
    context.user_data["state"] = "admin_manual_select_tariff"
    
    keyboard = [
        [KeyboardButton("💰 490 грн/год")],
        [KeyboardButton("💰 590 грн/год")],
        [KeyboardButton("🔙 Назад")]
    ]
    
    await update.message.reply_text(
        "💰 *Крок 3/7: Тариф*\n\nОберіть тариф:",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        parse_mode="Markdown"
    )

async def handle_admin_manual_select_tariff(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "🔙 Назад":
        context.user_data["state"] = "admin_manual_enter_name"
        await update.message.reply_text(
            "📝 Введіть ім'я та прізвище:",
            reply_markup=ReplyKeyboardMarkup([[KeyboardButton("🔙 Назад")]], resize_keyboard=True)
        )
        return
    
    tariff = int(text.split()[1])
    context.user_data["admin_booking"]["tariff"] = tariff
    context.user_data["state"] = "admin_manual_select_transmission"
    
    keyboard = [
        [KeyboardButton("🚗 Автомат")],
        [KeyboardButton("🚙 Механіка")],
        [KeyboardButton("🔙 Назад")]
    ]
    
    await update.message.reply_text(
        "🚗 *Крок 4/7: Коробка передач*\n\nОберіть тип:",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        parse_mode="Markdown"
    )

async def handle_admin_manual_select_transmission(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "🔙 Назад":
        context.user_data["state"] = "admin_manual_select_tariff"
        keyboard = [
            [KeyboardButton("💰 490 грн/год")],
            [KeyboardButton("💰 590 грн/год")],
            [KeyboardButton("🔙 Назад")]
        ]
        await update.message.reply_text(
            "💰 Оберіть тариф:",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return
    
    if text not in ["🚗 Автомат", "🚙 Механіка"]:
        await update.message.reply_text("⚠️ Оберіть коробку передач із меню.")
        return
    
    transmission = "Автомат" if text == "🚗 Автомат" else "Механіка"
    context.user_data["admin_booking"]["transmission"] = transmission
    
    instructors = get_instructors_by_transmission(transmission)
    if not instructors:
        await update.message.reply_text("😔 Немає інструкторів для цього типу.")
        return
    
    keyboard = []
    for instructor in instructors:
        rating = get_instructor_rating(instructor)
        if rating > 0:
            stars = "⭐" * int(rating)
            keyboard.append([KeyboardButton(f"{instructor} {stars} ({rating:.1f})")])
        else:
            keyboard.append([KeyboardButton(f"{instructor} 🆕")])
    keyboard.append([KeyboardButton("🔙 Назад")])
    
    context.user_data["state"] = "admin_manual_select_instructor"
    await update.message.reply_text(
        "👨‍🏫 *Крок 5/7: Інструктор*\n\nОберіть інструктора:",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        parse_mode="Markdown"
    )

async def handle_admin_manual_select_instructor(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "🔙 Назад":
        if context.user_data["admin_booking"].get("existing_student"):
            context.user_data["state"] = "admin_manual_confirm_student"
            await update.message.reply_text("Підтвердіть дані учня:")
        else:
            context.user_data["state"] = "admin_manual_select_transmission"
            keyboard = [
                [KeyboardButton("🚗 Автомат")],
                [KeyboardButton("🚙 Механіка")],
                [KeyboardButton("🔙 Назад")]
            ]
            await update.message.reply_text(
                "🚗 Оберіть коробку:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
        return
    
    instructor_name = text.split(" ⭐")[0].split(" 🆕")[0]
    context.user_data["admin_booking"]["instructor"] = instructor_name
    
    dates = get_next_dates(30, instructor_name)
    
    if not dates:
        await update.message.reply_text(
            f"😔 У інструктора {instructor_name} всі години зайняті на найближчі 30 днів\n\n"
            f"💡 Оберіть іншого інструктора або спробуйте пізніше."
        )
        return
    
    keyboard = []
    for date in dates:
        keyboard.append([KeyboardButton(date)])
    keyboard.append([KeyboardButton("🔙 Назад")])
    
    context.user_data["state"] = "admin_manual_select_date"
    await update.message.reply_text(
        "📅 *Крок 6/7: Дата*\n\nОберіть дату:",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        parse_mode="Markdown"
    )

async def handle_admin_manual_select_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "🔙 Назад":
        context.user_data["state"] = "admin_manual_select_instructor"
        await update.message.reply_text("👨‍🏫 Оберіть іншого інструктора:")
        return
    
    try:
        parts = text.split()
        
        # Шукаємо частину що містить крапку (дата формату ДД.ММ або ДД.ММ.РРРР)
        date_part = None
        for part in parts:
            clean = part.replace("(", "").replace(")", "")
            if clean.count('.') >= 1 and any(c.isdigit() for c in clean):
                date_part = clean
                break
        
        if not date_part:
            await update.message.reply_text("❌ Помилка парсингу дати. Оберіть зі списку.")
            return
        
        if date_part.count('.') == 1:
            day, month = date_part.split('.')
            year = datetime.now().year
            if int(month) < datetime.now().month:
                year += 1
            date_str = f"{day.zfill(2)}.{month.zfill(2)}.{year}"
        else:
            date_str = date_part
        
        context.user_data["admin_booking"]["date"] = date_str
        
        instructor = context.user_data["admin_booking"]["instructor"]
        free_slots = get_available_time_slots(instructor, date_str)
        
        if not free_slots:
            await update.message.reply_text("❌ Немає вільних часів. Оберіть іншу дату.")
            return
        
        keyboard = []
        for i in range(0, len(free_slots), 3):
            row = []
            for j in range(3):
                if i + j < len(free_slots):
                    row.append(KeyboardButton(free_slots[i + j]))
            keyboard.append(row)
        keyboard.append([KeyboardButton("🔙 Назад")])
        
        context.user_data["state"] = "admin_manual_select_time"
        await update.message.reply_text(
            f"🕐 *Крок 7/7: Час*\n\n"
            f"Оберіть час заняття:",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
            parse_mode="Markdown"
        )
    except Exception as e:
        logger.error(f"Error in handle_admin_manual_select_date: {e}", exc_info=True)
        await update.message.reply_text("❌ Помилка парсингу дати. Оберіть зі списку.")

async def handle_admin_manual_select_time(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "🔙 Назад":
        context.user_data["state"] = "admin_manual_select_date"
        await update.message.reply_text("📅 Оберіть іншу дату:")
        return
    
    if not re.match(r'^([0-1][0-9]|2[0-3]):[0-5][0-9]$', text):
        await update.message.reply_text("⚠️ Оберіть час з кнопок нижче.")
        return
    
    context.user_data["admin_booking"]["time"] = text
    context.user_data["state"] = "admin_manual_select_duration"
    
    keyboard = [
        [KeyboardButton("1 година")],
        [KeyboardButton("1.5 години")],
        [KeyboardButton("2 години")],
        [KeyboardButton("🔙 Назад")]
    ]
    
    await update.message.reply_text(
        "⏱ *Крок 8/8: Тривалість*\n\nОберіть тривалість:",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        parse_mode="Markdown"
    )

async def handle_admin_manual_select_duration(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "🔙 Назад":
        context.user_data["state"] = "admin_manual_select_time"
        await update.message.reply_text("🕐 Оберіть інший час:")
        return
    
    if text not in ["1 година", "1.5 години", "2 години"]:
        await update.message.reply_text("⚠️ Оберіть тривалість із меню.")
        return
    
    context.user_data["admin_booking"]["duration"] = text
    
    booking = context.user_data["admin_booking"]
    tariff = booking["tariff"]
    
    if "2" in text:
        price = tariff * 2
    elif "1.5" in text:
        price = tariff * 1.5
    else:
        price = tariff
    
    keyboard = [
        [KeyboardButton("✅ Підтвердити")],
        [KeyboardButton("🔙 Скасувати")]
    ]
    
    context.user_data["state"] = "admin_manual_confirm"
    await update.message.reply_text(
        f"📋 *Підтвердження запису*\n\n"
        f"👤 Учень: {booking['name']}\n"
        f"📱 Телефон: {booking['phone']}\n"
        f"👨‍🏫 Інструктор: {booking['instructor']}\n"
        f"📅 Дата: {booking['date']}\n"
        f"🕐 Час: {booking['time']}\n"
        f"⏱ Тривалість: {text}\n"
        f"💰 Вартість: {price:.0f} грн\n\n"
        f"{'📱 Учня буде сповіщено автоматично ✅' if booking.get('student_telegram_id') else '⚠️ Учень НЕ отримає повідомлення — зателефонуйте самостійно!'}",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        parse_mode="Markdown"
    )

async def handle_admin_manual_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "🔙 Скасувати":
        await show_admin_panel(update, context)
        return
    
    if text != "✅ Підтвердити":
        return
    
    booking = context.user_data["admin_booking"]
    
    instructor_data = get_instructor_by_name(booking["instructor"])
    if not instructor_data:
        await update.message.reply_text("❌ Помилка: інструктор не знайдений.")
        return
    
    instructor_id = instructor_data[0]
    student_telegram_id = booking.get("student_telegram_id")
    
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO lessons 
                (student_name, student_phone, student_tariff, instructor_id, date, time, duration, status, student_telegram_id, booking_comment)
                VALUES (?, ?, ?, ?, ?, ?, ?, 'active', ?, 'Запис адміном')
            """, (
                booking["name"],
                booking["phone"],
                booking["tariff"],
                instructor_id,
                booking["date"],
                booking["time"],
                booking["duration"],
                student_telegram_id
            ))
            conn.commit()
        
        if student_telegram_id:
            try:
                await context.bot.send_message(
                    chat_id=student_telegram_id,
                    text=(
                        f"✅ *Вас записано на заняття!*\n\n"
                        f"👨‍🏫 Інструктор: {booking['instructor']}\n"
                        f"📅 Дата: {booking['date']}\n"
                        f"🕐 Час: {booking['time']}\n"
                        f"⏱ Тривалість: {booking['duration']}\n\n"
                        f"Гарного навчання! 🚗"
                    ),
                    parse_mode="Markdown"
                )
                notify_status = "📱 Учню надіслано сповіщення ✅"
            except Exception as e:
                logger.error(f"Не вдалось надіслати повідомлення учню {student_telegram_id}: {e}")
                notify_status = f"⚠️ Не вдалось сповістити учня — зателефонуйте: {booking['phone']}"
        else:
            notify_status = f"📞 Зателефонуйте учню: {booking['phone']}"
        
        await update.message.reply_text(
            f"✅ *Запис створено!*\n\n"
            f"📋 Деталі:\n"
            f"{booking['name']} → {booking['instructor']}\n"
            f"{booking['date']} {booking['time']} ({booking['duration']})\n\n"
            f"{notify_status}",
            parse_mode="Markdown"
        )
        
        await show_admin_panel(update, context)
        
    except Exception as e:
        logger.error(f"Error creating manual booking: {e}")
        await update.message.reply_text("❌ Помилка при створенні запису.")

# ======================= STUDENT FUNCTIONS =======================
async def show_student_lessons(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT l.date, l.time, l.duration, i.name, i.phone, l.status, l.booking_comment
                FROM lessons l
                JOIN instructors i ON l.instructor_id = i.id
                WHERE l.student_telegram_id = ? AND l.status = 'active'
                ORDER BY l.date, l.time
                LIMIT 10
            """, (user_id,))
            
            lessons = cursor.fetchall()
        
        if not lessons:
            await update.message.reply_text("📋 У вас поки немає записів на заняття.")
            return
        
        text = "📖 Ваші записи:\n\n"
        
        for date, time, duration, instructor_name, instructor_phone, status, booking_comment in lessons:
            text += f"📅 {date} о {time} ({duration})\n"
            text += f"👨‍🏫 {instructor_name} | 📱 {instructor_phone}\n"
            if booking_comment:
                text += f"💬 Ваш коментар: \"{booking_comment}\"\n"
            text += "\n"
        
        await update.message.reply_text(text)
        
    except Exception as e:
        logger.error(f"Error in show_student_lessons: {e}", exc_info=True)
        await update.message.reply_text("❌ Помилка завантаження записів.")

async def show_student_statistics(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    
    try:
        from datetime import datetime, timedelta
        
        now = datetime.now(TZ)
        today_str = now.strftime("%d.%m.%Y")
        
        with get_db() as conn:
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT COUNT(*), 
                       SUM(CASE 
                           WHEN duration LIKE '%2%' THEN 2
                           WHEN duration LIKE '%1.5%' THEN 1.5
                           ELSE 1
                       END),
                       SUM(CASE 
                           WHEN duration LIKE '%2%' THEN student_tariff * 2
                           ELSE student_tariff
                       END)
                FROM lessons
                WHERE student_telegram_id = ? 
                AND status = 'active'
                AND date >= ?
            """, (user_id, today_str))
            
            planned = cursor.fetchone()
            planned_count = planned[0] or 0
            planned_hours = planned[1] or 0
            planned_cost = planned[2] or 0
            
            cursor.execute("""
                SELECT COUNT(*), 
                       SUM(CASE 
                           WHEN duration LIKE '%2%' THEN 2
                           WHEN duration LIKE '%1.5%' THEN 1.5
                           ELSE 1
                       END),
                       SUM(CASE 
                           WHEN duration LIKE '%2%' THEN student_tariff * 2
                           ELSE student_tariff
                       END)
                FROM lessons
                WHERE student_telegram_id = ? 
                AND status = 'completed'
            """, (user_id,))
            
            completed = cursor.fetchone()
            completed_count = completed[0] or 0
            completed_hours = completed[1] or 0
            completed_cost = completed[2] or 0
            
            cursor.execute("""
                SELECT AVG(rating), COUNT(rating)
                FROM lessons
                WHERE student_telegram_id = ?
                AND status = 'completed'
                AND rating IS NOT NULL
            """, (user_id,))
            
            rating_data = cursor.fetchone()
            avg_rating = rating_data[0] or 0
            rated_lessons = rating_data[1] or 0
            
            cursor.execute("""
                SELECT i.name, COUNT(*)
                FROM lessons l
                JOIN instructors i ON l.instructor_id = i.id
                WHERE l.student_telegram_id = ?
                AND l.status = 'completed'
                GROUP BY i.name
                ORDER BY COUNT(*) DESC
            """, (user_id,))
            
            instructors = cursor.fetchall()
        
        text = "📊 Статистика\n\n"
        
        text += "▶️ ЗАПЛАНОВАНО\n"
        if planned_count > 0:
            text += f"   {planned_count} {'урок' if planned_count == 1 else 'уроки' if planned_count < 5 else 'уроків'} "
            text += f"({planned_hours:.1f} год) → {planned_cost:,.0f} грн\n\n"
        else:
            text += "   Немає запланованих уроків\n\n"
        
        text += "✅ ЗАВЕРШЕНО\n"
        if completed_count > 0:
            text += f"   {completed_count} {'урок' if completed_count == 1 else 'уроки' if completed_count < 5 else 'уроків'} "
            text += f"({completed_hours:.1f} год) → {completed_cost:,.0f} грн\n\n"
        else:
            text += "   Поки немає завершених уроків\n\n"
        
        if rated_lessons > 0:
            text += "📈 ПРОГРЕС\n"
            text += f"   ⭐ Середня оцінка: {avg_rating:.1f}/5 (за {rated_lessons} {'урок' if rated_lessons == 1 else 'уроки' if rated_lessons < 5 else 'уроків'})\n\n"
        
        if instructors:
            text += "👨‍🏫 ІНСТРУКТОРИ\n"
            instructor_names = []
            for name, count in instructors:
                short_name = name.split()[0]
                instructor_names.append(f"{short_name}: {count}")
            text += f"   {' | '.join(instructor_names)}\n"
        
        await update.message.reply_text(text)
        
    except Exception as e:
        logger.error(f"Error in show_student_statistics: {e}", exc_info=True)
        await update.message.reply_text("❌ Помилка завантаження статистики.")

# ======================= CANCEL LESSON FUNCTIONS =======================
async def show_lessons_to_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    
    try:
        now = datetime.now(TZ)
        
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT l.id, l.date, l.time, l.duration, i.name
                FROM lessons l
                JOIN instructors i ON l.instructor_id = i.id
                WHERE l.student_telegram_id = ? AND l.status = 'active'
                ORDER BY l.date, l.time
                LIMIT 10
            """, (user_id,))
            
            lessons = cursor.fetchall()
        
        if not lessons:
            await update.message.reply_text("📋 У вас немає активних записів на заняття.")
            return
        
        cancelable_lessons = []
        
        for lesson_id, date, time, duration, instructor_name in lessons:
            try:
                lesson_datetime = datetime.strptime(f"{date} {time}", "%d.%m.%Y %H:%M")
                lesson_datetime = TZ.localize(lesson_datetime)
                
                hours_until = (lesson_datetime - now).total_seconds() / 3600
                
                if hours_until >= 12:
                    cancelable_lessons.append((lesson_id, date, time, duration, instructor_name, hours_until))
            except Exception as e:
                logger.error(f"Error parsing lesson time: {e}")
                continue
        
        if not cancelable_lessons:
            await update.message.reply_text(
                "⚠️ Немає уроків які можна скасувати\n\n"
                "Скасування можливе мінімум за 12 годин до уроку."
            )
            return
        
        context.user_data["cancelable_lessons"] = cancelable_lessons
        context.user_data["state"] = "cancel_lesson_select"
        
        text = "❌ *Скасування запису*\n\n"
        text += "Оберіть урок для скасування:\n\n"
        
        keyboard = []
        
        for i, (lesson_id, date, time, duration, instructor_name, hours_until) in enumerate(cancelable_lessons, 1):
            text += f"{i}. {date} {time} ({duration})\n"
            text += f"   👨‍🏫 {instructor_name}\n"
            text += f"   ⏰ Залишилось {int(hours_until)} год\n\n"
            keyboard.append([KeyboardButton(f"{i}")])
        
        keyboard.append([KeyboardButton("🔙 Назад")])
        
        await update.message.reply_text(
            text,
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
            parse_mode="Markdown"
        )
        
    except Exception as e:
        logger.error(f"Error in show_lessons_to_cancel: {e}", exc_info=True)
        await update.message.reply_text("❌ Помилка завантаження записів.")

async def handle_cancel_lesson(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "🔙 Назад":
        context.user_data.clear()
        await start(update, context)
        return
    
    try:
        lesson_index = int(text) - 1
        lessons = context.user_data.get("cancelable_lessons", [])
        
        if lesson_index < 0 or lesson_index >= len(lessons):
            await update.message.reply_text("⚠️ Невірний номер. Спробуйте ще раз:")
            return
        
        selected = lessons[lesson_index]
        lesson_id, date, time, duration, instructor_name, hours_until = selected
        
        context.user_data["cancel_lesson_id"] = lesson_id
        context.user_data["cancel_lesson_date"] = date
        context.user_data["cancel_lesson_time"] = time
        context.user_data["cancel_lesson_instructor"] = instructor_name
        context.user_data["state"] = "cancel_lesson_confirm"
        
        keyboard = [
            [KeyboardButton("✅ Так, скасувати")],
            [KeyboardButton("🔙 Ні, залишити")]
        ]
        
        await update.message.reply_text(
            f"⚠️ *Підтвердіть скасування*\n\n"
            f"📅 Дата: {date}\n"
            f"🕐 Час: {time}\n"
            f"⏱ Тривалість: {duration}\n"
            f"👨‍🏫 Інструктор: {instructor_name}\n\n"
            f"Скасувати урок?",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
            parse_mode="Markdown"
        )
        
    except ValueError:
        await update.message.reply_text("⚠️ Введіть номер уроку:")
        return
    except Exception as e:
        logger.error(f"Error in handle_cancel_lesson: {e}", exc_info=True)
        await update.message.reply_text("❌ Помилка.")

async def handle_cancel_confirmation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "🔙 Ні, залишити":
        context.user_data.clear()
        await update.message.reply_text("✅ Запис залишено без змін.")
        await start(update, context)
        return
    
    if text != "✅ Так, скасувати":
        await update.message.reply_text("⚠️ Оберіть дію з меню:")
        return
    
    try:
        lesson_id = context.user_data.get("cancel_lesson_id")
        date = context.user_data.get("cancel_lesson_date")
        time = context.user_data.get("cancel_lesson_time")
        instructor_name = context.user_data.get("cancel_lesson_instructor")
        
        user_id = update.message.from_user.id
        
        with get_db() as conn:
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT student_name, student_phone, student_tariff, duration, instructor_id
                FROM lessons
                WHERE id = ?
            """, (lesson_id,))
            
            lesson_data = cursor.fetchone()
            
            if not lesson_data:
                await update.message.reply_text("❌ Урок не знайдено.")
                return
            
            student_name, student_phone, student_tariff, duration, instructor_id = lesson_data
            
            cursor.execute("""
                UPDATE lessons
                SET status = 'cancelled',
                    cancelled_by = 'student',
                    cancelled_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (lesson_id,))
            
            cursor.execute("SELECT telegram_id FROM instructors WHERE id = ?", (instructor_id,))
            instructor_telegram_id = cursor.fetchone()[0]
            
            conn.commit()
        
        await update.message.reply_text(
            f"✅ *Урок скасовано!*\n\n"
            f"📅 {date} {time}\n"
            f"👨‍🏫 {instructor_name}",
            parse_mode="Markdown"
        )
        
        if instructor_telegram_id:
            try:
                if student_tariff and "2" in duration:
                    price = student_tariff * 2
                elif student_tariff:
                    price = student_tariff
                else:
                    price = PRICES.get(duration, 420)
                
                await context.bot.send_message(
                    chat_id=instructor_telegram_id,
                    text=f"🔔 *Урок скасовано учнем*\n\n"
                         f"👤 Учень: {student_name}\n"
                         f"📱 Телефон: {student_phone}\n"
                         f"📅 Дата: {date}\n"
                         f"🕐 Час: {time}\n"
                         f"⏱ Тривалість: {duration}\n"
                         f"💰 Сума: {price:.0f} грн",
                    parse_mode="Markdown"
                )
            except Exception as e:
                logger.error(f"Failed to notify instructor: {e}")
        
        context.user_data.clear()
        await start(update, context)
        
    except Exception as e:
        logger.error(f"Error in handle_cancel_confirmation: {e}", exc_info=True)
        await update.message.reply_text("❌ Помилка скасування.")


async def save_lesson(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Зберегти заняття в БД"""
    try:
        instructor_name = context.user_data["instructor"]
        date = context.user_data["date"]
        time = context.user_data["time"]
        duration = context.user_data["duration"]
        student_name = context.user_data.get("student_name", "")
        student_phone = context.user_data.get("student_phone", "")
        student_telegram_id = update.message.from_user.id
        student_tariff = context.user_data.get("student_tariff", 0)
        
        instructor_data = get_instructor_by_name(instructor_name)
        if not instructor_data:
            await update.message.reply_text("❌ Помилка: інструктор не знайдений.")
            return
        
        instructor_id, instructor_telegram_id = instructor_data
        
        start_hour = int(time.split(':')[0])
        if "2" in duration:
            lesson_hours = 2
        elif "1.5" in duration:
            lesson_hours = 1.5
        else:
            lesson_hours = 1
        
        end_hour = start_hour + lesson_hours
        
        with get_db() as conn:
            conn.execute("BEGIN IMMEDIATE")
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT i.name, l.time, l.duration
                FROM lessons l
                JOIN instructors i ON l.instructor_id = i.id
                WHERE l.student_telegram_id = ? AND l.date = ? AND l.status = 'active'
            """, (student_telegram_id, date))
            
            existing_lessons = cursor.fetchall()
            
            for existing_instructor, existing_time, existing_duration in existing_lessons:
                existing_start = int(existing_time.split(':')[0])
                if "2" in existing_duration:
                    existing_hours = 2
                elif "1.5" in existing_duration:
                    existing_hours = 1.5
                else:
                    existing_hours = 1
                existing_end = existing_start + existing_hours
                
                if not (end_hour <= existing_start or start_hour >= existing_end):
                    await update.message.reply_text(
                        f"❌ *Не можна записатись!*\n\n"
                        f"У вас вже є урок в цей час:\n"
                        f"👨‍🏫 {existing_instructor}\n"
                        f"📅 {date}\n"
                        f"🕐 {existing_time} ({existing_duration})\n\n"
                        f"Оберіть інший час.",
                        parse_mode="Markdown"
                    )
                    return
            
            cursor.execute("""
                SELECT SUM(
                    CASE 
                        WHEN duration LIKE '%2%' THEN 2
                        WHEN duration LIKE '%1.5%' THEN 1.5
                        ELSE 1
                    END
                )
                FROM lessons
                WHERE student_telegram_id = ? AND date = ? AND status = 'active'
            """, (student_telegram_id, date))
            
            total_hours_today = cursor.fetchone()[0] or 0
            
            if total_hours_today + lesson_hours > 2:
                await update.message.reply_text(
                    f"❌ *Ліміт перевищено!*\n\n"
                    f"Ви вже маєте *{total_hours_today:.1f} год* на {date}\n"
                    f"Максимум: *2 години на день*\n\n"
                    f"Залишилось: *{2 - total_hours_today:.1f} год*",
                    parse_mode="Markdown"
                )
                return
            
            from datetime import datetime, timedelta
            date_obj = datetime.strptime(date, "%d.%m.%Y")
            week_start = date_obj - timedelta(days=date_obj.weekday())
            week_end = week_start + timedelta(days=6)
            
            cursor.execute("""
                SELECT SUM(
                    CASE 
                        WHEN duration LIKE '%2%' THEN 2
                        WHEN duration LIKE '%1.5%' THEN 1.5
                        ELSE 1
                    END
                )
                FROM lessons
                WHERE student_telegram_id = ? 
                AND substr(date, 7, 4) || '-' || substr(date, 4, 2) || '-' || substr(date, 1, 2) 
                    BETWEEN ? AND ?
                AND status = 'active'
            """, (student_telegram_id, 
                  week_start.strftime("%Y-%m-%d"), 
                  week_end.strftime("%Y-%m-%d")))
            
            total_hours_week = cursor.fetchone()[0] or 0
            
            if total_hours_week + lesson_hours > 6:
                await update.message.reply_text(
                    f"❌ *Ліміт перевищено!*\n\n"
                    f"Ви вже маєте *{total_hours_week:.1f} год* цього тижня\n"
                    f"Максимум: *6 годин на тиждень*\n\n"
                    f"Залишилось: *{6 - total_hours_week:.1f} год*",
                    parse_mode="Markdown"
                )
                return
            
            cursor.execute("""
                SELECT student_name, student_telegram_id, time, duration
                FROM lessons
                WHERE instructor_id = ? AND date = ? AND status = 'active'
            """, (instructor_id, date))
            
            instructor_lessons = cursor.fetchall()
            
            for other_student_name, other_student_id, other_time, other_duration in instructor_lessons:
                if other_student_id == student_telegram_id:
                    continue
                
                other_start = int(other_time.split(':')[0])
                if "2" in other_duration:
                    other_hours = 2
                elif "1.5" in other_duration:
                    other_hours = 1.5
                else:
                    other_hours = 1
                other_end = other_start + other_hours
                
                if not (end_hour <= other_start or start_hour >= other_end):
                    await update.message.reply_text(
                        f"❌ *Інструктор зайнятий!*\n\n"
                        f"На цей час вже записаний інший учень:\n"
                        f"👤 {other_student_name}\n"
                        f"📅 {date}\n"
                        f"🕐 {other_time} ({other_duration})\n\n"
                        f"Оберіть інший час або дату.",
                        parse_mode="Markdown"
                    )
                    return
            
            booking_comment = context.user_data.get("booking_comment", "")
            
            cursor.execute("""
                INSERT INTO lessons 
                (instructor_id, student_name, student_telegram_id, student_phone, student_tariff, date, time, duration, status, booking_comment)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'active', ?)
            """, (instructor_id, student_name, student_telegram_id, student_phone, student_tariff, date, time, duration, booking_comment))
            conn.commit()
        
        await update.message.reply_text(
            f"✅ *Заняття заброньовано!*\n\n"
            f"👨‍🏫 Інструктор: {instructor_name}\n"
            f"📅 Дата: {date}\n"
            f"🕐 Час: {time}\n"
            f"⏱ Тривалість: {duration}",
            parse_mode="Markdown"
        )
        
        if student_tariff > 0:
            if "2" in duration:
                price = student_tariff * 2
            else:
                price = student_tariff
        else:
            price = PRICES.get(duration, 420)
        
        booking_comment = context.user_data.get("booking_comment", "")
        
        if instructor_telegram_id:
            try:
                message_text = (
                    f"🔔 *Новий запис!*\n\n"
                    f"👤 Учень: {student_name}\n"
                    f"📱 Телефон: {student_phone}\n"
                    f"📅 Дата: {date}\n"
                    f"🕐 Час: {time}\n"
                    f"⏱ Тривалість: {duration}\n"
                    f"💰 Вартість: *{price:.0f} грн*"
                )
                
                if booking_comment:
                    message_text += f"\n\n💬 Коментар учня:\n\"{booking_comment}\""
                
                await context.bot.send_message(
                    chat_id=instructor_telegram_id,
                    text=message_text,
                    parse_mode="Markdown"
                )
            except Exception as e:
                logger.error(f"Не вдалося надіслати повідомлення інструктору: {e}")
        
        await start(update, context)
        
    except Exception as e:
        logger.error(f"Error in save_lesson: {e}", exc_info=True)
        await update.message.reply_text("❌ Помилка збереження запису.")

# ======================= CALLBACKS =======================
async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    try:
        if query.data.startswith("unblock_"):
            block_id = int(query.data.split("_")[1])
            await handle_unblock_callback(query, context, block_id)
            
    except Exception as e:
        logger.error(f"Error in handle_callback: {e}", exc_info=True)
        await query.edit_message_text("❌ Помилка.")

async def handle_unblock_callback(query, context, block_id):
    try:
        from database import remove_schedule_block
        
        if remove_schedule_block(block_id):
            await query.edit_message_text("✅ Час розблоковано!")
        else:
            await query.edit_message_text("❌ Помилка розблокування.")
            
    except Exception as e:
        logger.error(f"Error in handle_unblock_callback: {e}", exc_info=True)
        await query.edit_message_text("❌ Помилка.")

# ======================= REMINDERS =======================
async def send_reminders(context: ContextTypes.DEFAULT_TYPE):
    try:
        now = datetime.now(TZ)
        logger.info(f"🔔 send_reminders запущено! Зараз: {now.strftime('%d.%m.%Y %H:%M')}")
        
        with get_db() as conn:
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT l.id, l.student_telegram_id, i.name, l.date, l.time
                FROM lessons l
                JOIN instructors i ON l.instructor_id = i.id
                WHERE l.status = 'active' 
                AND l.reminder_24h_sent = 0
            """)
            
            all_lessons = cursor.fetchall()
            lessons_24h = []
            
            for lesson_id, student_id, instructor, date_str, time_str in all_lessons:
                try:
                    lesson_datetime = datetime.strptime(f"{date_str} {time_str}", "%d.%m.%Y %H:%M")
                    lesson_datetime = TZ.localize(lesson_datetime)
                    
                    time_diff = (lesson_datetime - now).total_seconds() / 3600
                    
                    if 23.5 <= time_diff <= 24.5:
                        lessons_24h.append((lesson_id, student_id, instructor, date_str, time_str))
                except Exception as e:
                    logger.error(f"Error parsing lesson date {date_str} {time_str}: {e}")
            
            for lesson_id, student_id, instructor, date, time in lessons_24h:
                try:
                    await context.bot.send_message(
                        chat_id=student_id,
                        text=f"⏰ *Нагадування!*\n\nУ вас заняття завтра:\n"
                             f"👨‍🏫 {instructor}\n📅 {date}\n🕐 {time}",
                        parse_mode="Markdown",
                        reply_markup=ReplyKeyboardRemove()
                    )
                    
                    cursor.execute("UPDATE lessons SET reminder_24h_sent = 1 WHERE id = ?", (lesson_id,))
                    conn.commit()
                except Exception as e:
                    logger.error(f"Failed to send 24h reminder: {e}")
            
            cursor.execute("""
                SELECT l.id, l.student_telegram_id, i.name, l.date, l.time
                FROM lessons l
                JOIN instructors i ON l.instructor_id = i.id
                WHERE l.status = 'active' 
                AND l.reminder_2h_sent = 0
            """)
            
            all_lessons_2h = cursor.fetchall()
            lessons_2h = []
            
            for lesson_id, student_id, instructor, date_str, time_str in all_lessons_2h:
                try:
                    lesson_datetime = datetime.strptime(f"{date_str} {time_str}", "%d.%m.%Y %H:%M")
                    lesson_datetime = TZ.localize(lesson_datetime)
                    
                    time_diff = (lesson_datetime - now).total_seconds() / 3600
                    
                    if 1.5 <= time_diff <= 2.5:
                        lessons_2h.append((lesson_id, student_id, instructor, date_str, time_str))
                except Exception as e:
                    logger.error(f"Error parsing lesson date {date_str} {time_str}: {e}")
            
            for lesson_id, student_id, instructor, date, time in lessons_2h:
                try:
                    await context.bot.send_message(
                        chat_id=student_id,
                        text=f"🔔 *Нагадування!*\n\nУ вас заняття через 2 години:\n"
                             f"👨‍🏫 {instructor}\n📅 {date}\n🕐 {time}\n\n"
                             f"⏰ Не забудьте підготуватися!",
                        parse_mode="Markdown",
                        reply_markup=ReplyKeyboardRemove()
                    )
                    
                    cursor.execute("UPDATE lessons SET reminder_2h_sent = 1 WHERE id = ?", (lesson_id,))
                    conn.commit()
                except Exception as e:
                    logger.error(f"Failed to send 2h reminder: {e}")
        
        logger.info("✅ Reminders sent successfully")
        
    except Exception as e:
        logger.error(f"Error in send_reminders: {e}", exc_info=True)

async def send_rating_request_to_student(context, student_tg_id, lesson_id, date, time, instructor_name):
    try:
        keyboard = [
            [KeyboardButton("⭐"), KeyboardButton("⭐⭐"), KeyboardButton("⭐⭐⭐")],
            [KeyboardButton("⭐⭐⭐⭐"), KeyboardButton("⭐⭐⭐⭐⭐")],
            [KeyboardButton("⏭️ Пропустити")]
        ]
        
        await context.bot.send_message(
            chat_id=student_tg_id,
            text=f"✅ *Урок завершено!*\n\n"
                 f"📅 {date} {time}\n"
                 f"👨‍🏫 {instructor_name}\n\n"
                 f"⭐ Оцініть інструктора:",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
            parse_mode="Markdown"
        )
        
        context.bot_data[f"rating_lesson_{student_tg_id}"] = {
            'lesson_id': lesson_id,
            'instructor_name': instructor_name,
            'date': date,
            'time': time
        }
        
    except Exception as e:
        logger.error(f"Error sending rating request: {e}", exc_info=True)

async def check_completed_lessons(context: ContextTypes.DEFAULT_TYPE):
    try:
        now = datetime.now(TZ)
        
        with get_db() as conn:
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT l.id, l.date, l.time, l.student_telegram_id, 
                       l.instructor_id, i.name as instructor_name
                FROM lessons l
                JOIN instructors i ON l.instructor_id = i.id
                WHERE l.status = 'active'
            """)
            
            lessons_to_complete = []
            
            for lesson_id, date_str, time_str, student_tg_id, instructor_id, instructor_name in cursor.fetchall():
                try:
                    lesson_datetime = datetime.strptime(f"{date_str} {time_str}", "%d.%m.%Y %H:%M")
                    lesson_datetime = TZ.localize(lesson_datetime)
                    
                    if lesson_datetime < now:
                        lessons_to_complete.append({
                            'id': lesson_id,
                            'date': date_str,
                            'time': time_str,
                            'student_tg_id': student_tg_id,
                            'instructor_id': instructor_id,
                            'instructor_name': instructor_name
                        })
                except Exception as e:
                    logger.error(f"Error parsing lesson date {date_str} {time_str}: {e}")
            
            for lesson in lessons_to_complete:
                cursor.execute("""
                    UPDATE lessons
                    SET status = 'completed', completed_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (lesson['id'],))
            
            conn.commit()
            
            if lessons_to_complete:
                logger.info(f"Completed {len(lessons_to_complete)} lessons")
                
                for lesson in lessons_to_complete:
                    try:
                        await send_rating_request_to_student(
                            context, 
                            lesson['student_tg_id'],
                            lesson['id'],
                            lesson['date'],
                            lesson['time'],
                            lesson['instructor_name']
                        )
                    except Exception as e:
                        logger.error(f"Failed to send rating request for lesson {lesson['id']}: {e}")
        
    except Exception as e:
        logger.error(f"Error in check_completed_lessons: {e}", exc_info=True)

# ======================= EXPORT WITH PERIOD SELECTION =======================
async def show_export_period_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [KeyboardButton("📊 За тиждень")],
        [KeyboardButton("📊 За місяць")],
        [KeyboardButton("📊 За весь час")],
        [KeyboardButton("📊 Свій період")],
        [KeyboardButton("🔙 Назад")]
    ]
    
    context.user_data["state"] = "export_period"
    
    await update.message.reply_text(
        "📥 *Експорт в Excel*\n\n"
        "Оберіть період для експорту:",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        parse_mode="Markdown"
    )

async def handle_export_period_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "🔙 Назад":
        await show_admin_panel(update, context)
        return
    
    from datetime import datetime, timedelta
    today = datetime.now().date()
    
    if text == "📊 За тиждень":
        date_from = (today - timedelta(days=7)).strftime("%d.%m.%Y")
        date_to = (today + timedelta(days=7)).strftime("%d.%m.%Y")
        period_name = "тиждень"
        
    elif text == "📊 За місяць":
        date_from = (today - timedelta(days=30)).strftime("%d.%m.%Y")
        date_to = (today + timedelta(days=30)).strftime("%d.%m.%Y")
        period_name = "місяць"
        
    elif text == "📊 За весь час":
        date_from = "01.01.2020"
        date_to = (today + timedelta(days=365)).strftime("%d.%m.%Y")
        period_name = "весь час"
        
    elif text == "📊 Свій період":
        context.user_data["state"] = "export_custom_period"
        
        keyboard = [[KeyboardButton("🔙 Назад")]]
        
        await update.message.reply_text(
            "📅 *Введіть період у форматі:*\n"
            "`ДД.ММ.РРРР - ДД.ММ.РРРР`\n\n"
            "Наприклад: `01.12.2025 - 31.12.2025`",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
            parse_mode="Markdown"
        )
        return
    else:
        await update.message.reply_text("⚠️ Оберіть період з меню.")
        return
    
    await export_to_excel_with_period(update, context, date_from, date_to, period_name)

async def handle_export_custom_period(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "🔙 Назад":
        await show_export_period_menu(update, context)
        return
    
    try:
        import re
        match = re.match(r'(\d{2}\.\d{2}\.\d{4})\s*-\s*(\d{2}\.\d{2}\.\d{4})', text)
        
        if not match:
            await update.message.reply_text(
                "⚠️ Невірний формат!\n\n"
                "Використовуйте: `ДД.ММ.РРРР - ДД.ММ.РРРР`\n"
                "Наприклад: `01.12.2025 - 31.12.2025`",
                parse_mode="Markdown"
            )
            return
        
        date_from = match.group(1)
        date_to = match.group(2)
        
        from datetime import datetime
        try:
            datetime.strptime(date_from, "%d.%m.%Y")
            datetime.strptime(date_to, "%d.%m.%Y")
        except ValueError:
            await update.message.reply_text("⚠️ Невірна дата! Перевірте формат.")
            return
        
        period_name = f"{date_from} - {date_to}"
        
        await export_to_excel_with_period(update, context, date_from, date_to, period_name)
        
    except Exception as e:
        logger.error(f"Error in handle_export_custom_period: {e}", exc_info=True)
        await update.message.reply_text("❌ Помилка обробки періоду.")

async def export_to_excel_with_period(update: Update, context: ContextTypes.DEFAULT_TYPE, date_from: str, date_to: str, period_name: str):
    user_id = update.message.from_user.id
    
    try:
        await update.message.reply_text("⏳ Генерую Excel файл... Зачекайте...")
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        wb = Workbook()
        
        ws1 = wb.active
        ws1.title = "Уроки"
        
        headers1 = ["ID", "Дата", "Час", "Інструктор", "Учень", "Телефон", "Тариф", "Тривалість", "Вартість", "Статус", "⭐ Оцінка учня", "💬 Коментар учня", "⭐ Оцінка інструктора", "💬 Коментар інструктора"]
        ws1.append(headers1)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        with get_db() as conn:
            cursor = conn.cursor()
            
            cursor.execute("PRAGMA table_info(lessons)")
            columns = {row[1] for row in cursor.fetchall()}
            has_instructor_rating = 'instructor_rating' in columns
            has_instructor_feedback = 'instructor_feedback' in columns
            
            if has_instructor_rating and has_instructor_feedback:
                cursor.execute("""
                    SELECT 
                        l.id, l.date, l.time,
                        i.name as instructor_name,
                        s.name as student_name,
                        s.phone as student_phone,
                        s.tariff,
                        l.duration,
                        CASE 
                            WHEN l.duration LIKE '%2%' THEN s.tariff * 2
                            WHEN l.duration LIKE '%1.5%' THEN s.tariff * 1.5
                            ELSE s.tariff * 1
                        END as earnings,
                        l.status, l.rating, l.feedback,
                        l.instructor_rating, l.instructor_feedback
                    FROM lessons l
                    LEFT JOIN instructors i ON l.instructor_id = i.id
                    LEFT JOIN students s ON l.student_telegram_id = s.telegram_id
                    ORDER BY l.date DESC, l.time DESC
                """)
            else:
                cursor.execute("""
                    SELECT 
                        l.id, l.date, l.time,
                        i.name as instructor_name,
                        s.name as student_name,
                        s.phone as student_phone,
                        s.tariff,
                        l.duration,
                        CASE 
                            WHEN l.duration LIKE '%2%' THEN s.tariff * 2
                            WHEN l.duration LIKE '%1.5%' THEN s.tariff * 1.5
                            ELSE s.tariff * 1
                        END as earnings,
                        l.status, l.rating, l.feedback,
                        NULL as instructor_rating,
                        NULL as instructor_feedback
                    FROM lessons l
                    LEFT JOIN instructors i ON l.instructor_id = i.id
                    LEFT JOIN students s ON l.student_telegram_id = s.telegram_id
                    ORDER BY l.date DESC, l.time DESC
                """)
            
            all_lessons = cursor.fetchall()
        
        from datetime import datetime as dt
        date_from_obj = dt.strptime(date_from, "%d.%m.%Y")
        date_to_obj = dt.strptime(date_to, "%d.%m.%Y")
        
        lessons = []
        for lesson in all_lessons:
            try:
                lesson_date = dt.strptime(lesson[1], "%d.%m.%Y")
                if date_from_obj <= lesson_date <= date_to_obj:
                    lessons.append(lesson)
            except (ValueError, TypeError):
                continue
        
        total_lessons = len(lessons)
        total_earnings = 0
        unique_students = set()
        
        for lesson in lessons:
            ws1.append(lesson)
            if lesson[8]:
                total_earnings += lesson[8]
            if lesson[4]:
                unique_students.add(lesson[4])
        
        for column in ws1.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws1.column_dimensions[column[0].column_letter].width = adjusted_width
        
        ws2 = wb.create_sheet(title="Учні")
        headers2 = ["Учень", "Телефон", "Тариф", "Уроків", "Годин", "Витрачено", "⭐ Середній рейтинг"]
        ws2.append(headers2)
        
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        students_stats = {}
        
        for lesson in lessons:
            student_name = lesson[4]
            student_phone = lesson[5]
            student_tariff = lesson[6]
            duration = lesson[7]
            earnings = lesson[8]
            instructor_rating = lesson[12]
            
            if not student_name or not student_tariff:
                continue
            
            if student_name not in students_stats:
                students_stats[student_name] = {
                    'phone': student_phone,
                    'tariff': student_tariff,
                    'lessons': 0,
                    'hours': 0,
                    'spent': 0,
                    'ratings': []
                }
            
            students_stats[student_name]['lessons'] += 1
            
            if "1.5" in duration:
                students_stats[student_name]['hours'] += 1.5
            elif "2" in duration:
                students_stats[student_name]['hours'] += 2
            else:
                students_stats[student_name]['hours'] += 1
            
            if earnings:
                students_stats[student_name]['spent'] += earnings
            
            if instructor_rating and instructor_rating > 0:
                students_stats[student_name]['ratings'].append(instructor_rating)
        
        for name, stats in sorted(students_stats.items(), key=lambda x: x[1]['lessons'], reverse=True):
            avg_rating = sum(stats['ratings']) / len(stats['ratings']) if stats['ratings'] else None
            if avg_rating:
                avg_rating = round(avg_rating, 1)
            
            ws2.append((
                name,
                stats['phone'],
                stats['tariff'],
                stats['lessons'],
                stats['hours'],
                stats['spent'],
                avg_rating if avg_rating else '-'
            ))
        
        for column in ws2.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws2.column_dimensions[column[0].column_letter].width = adjusted_width
        
        ws3 = wb.create_sheet(title="Інструктори")
        headers3 = ["Інструктор", "Тариф", "Уроків", "Годин", "Заробіток", "Рейтинг"]
        ws3.append(headers3)
        
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id, name, price_per_hour FROM instructors")
            all_instructors = {row[0]: (row[1], row[2]) for row in cursor.fetchall()}
        
        instructors_stats = {}
        
        for lesson in lessons:
            instructor_name = lesson[3]
            if not instructor_name:
                continue
            
            instructor_id = None
            instructor_price = 0
            for iid, (iname, iprice) in all_instructors.items():
                if iname == instructor_name:
                    instructor_id = iid
                    instructor_price = iprice
                    break
            
            if not instructor_id:
                continue
            
            if instructor_id not in instructors_stats:
                instructors_stats[instructor_id] = {
                    'name': instructor_name,
                    'price': instructor_price,
                    'lessons': 0,
                    'hours': 0,
                    'earnings': 0,
                    'ratings': []
                }
            
            instructors_stats[instructor_id]['lessons'] += 1
            
            duration = lesson[7]
            if "1.5" in duration:
                instructors_stats[instructor_id]['hours'] += 1.5
            elif "2" in duration:
                instructors_stats[instructor_id]['hours'] += 2
            else:
                instructors_stats[instructor_id]['hours'] += 1
            
            earnings = lesson[8]
            if earnings:
                instructors_stats[instructor_id]['earnings'] += earnings
            
            rating = lesson[10]
            if rating and rating > 0:
                instructors_stats[instructor_id]['ratings'].append(rating)
        
        for iid, stats in sorted(instructors_stats.items(), key=lambda x: x[1]['lessons'], reverse=True):
            avg_rating = sum(stats['ratings']) / len(stats['ratings']) if stats['ratings'] else 0
            ws3.append((
                stats['name'],
                stats['price'],
                stats['lessons'],
                stats['hours'],
                stats['earnings'],
                round(avg_rating, 1)
            ))
        
        for column in ws3.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws3.column_dimensions[column[0].column_letter].width = adjusted_width

        ws4 = wb.create_sheet(title="Заблоковані часи")
        headers4 = ["Інструктор", "Дата", "Час початку", "Час кінця", "Причина", "Створено"]
        ws4.append(headers4)
        
        for cell in ws4[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT 
                    i.name AS instructor_name,
                    sb.date, sb.time_start, sb.time_end, sb.reason, sb.created_at
                FROM schedule_blocks sb
                JOIN instructors i ON sb.instructor_id = i.id
                ORDER BY sb.date DESC, sb.time_start
            """)
            blocked_times = cursor.fetchall()
        
        if blocked_times:
            for block in blocked_times:
                instructor_name = block[0]
                date = block[1]
                time_start = block[2]
                time_end = block[3]
                reason = block[4] or "Не вказано"
                created_at = block[5]
                
                try:
                    date_obj = datetime.strptime(date, "%Y-%m-%d")
                    date_formatted = date_obj.strftime("%d.%m.%Y")
                    weekday = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Нд"][date_obj.weekday()]
                    date_display = f"{weekday} {date_formatted}"
                except:
                    date_display = date
                
                try:
                    created_obj = datetime.strptime(created_at, "%Y-%m-%d %H:%M:%S")
                    created_display = created_obj.strftime("%d.%m.%Y %H:%M")
                except:
                    created_display = created_at
                
                ws4.append([
                    instructor_name,
                    date_display,
                    time_start,
                    time_end,
                    reason,
                    created_display
                ])
        else:
            ws4.append(["Немає заблокованих часів", "", "", "", "", ""])
        
        for column in ws4.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws4.column_dimensions[column[0].column_letter].width = adjusted_width
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        filename = f"export_{period_name.replace(' ', '_').replace(':', '-')}.xlsx"
        
        await context.bot.send_document(
            chat_id=update.message.chat_id,
            document=excel_file,
            filename=filename,
            caption=f"📊 *Експорт завершено!*\n\n"
                   f"📅 Період: {period_name}\n"
                   f"📝 Уроків: {total_lessons}\n"
                   f"👥 Учнів: {len(unique_students)}\n"
                   f"💰 Загальний заробіток: {total_earnings:.0f} грн",
            parse_mode="Markdown"
        )
        
        logger.info(f"✅ Excel exported for period: {period_name}")
        
        await show_admin_panel(update, context)
        
    except Exception as e:
        logger.error(f"Error in export_to_excel_with_period: {e}", exc_info=True)
        await update.message.reply_text(
            "❌ Помилка експорту.\n\n"
            f"Деталі: {str(e)}"
        )
        await show_admin_panel(update, context)

# ======================= MAIN =======================
def main():
    try:
        os.environ["DB_NAME"] = DB_NAME
        
        logger.info("🚀 PRODUCTION ВЕРСІЯ БОТА")
        logger.info(f"🔑 TOKEN: {TOKEN[:20]}...")
        logger.info(f"💾 БД: {DB_NAME}")
        
        init_db()
        init_lessons_table()
        init_students_table()
        migrate_database()
        init_schedule_blocks_table()
        
        ensure_instructors_exist()

        from telegram.ext import JobQueue
        app = (
            ApplicationBuilder()
            .token(TOKEN)
            .build()
        )

        app.add_handler(CommandHandler("start", start))
        app.add_handler(CommandHandler("register490", register_490))
        app.add_handler(CommandHandler("register590", register_590))
        
        app.add_handler(CallbackQueryHandler(handle_callback))
        app.add_handler(MessageHandler(filters.TEXT & (~filters.COMMAND), handle_message))
        app.add_handler(MessageHandler(filters.CONTACT, handle_message))

        if app.job_queue:
            app.job_queue.run_repeating(send_reminders, interval=1800, first=10)
            app.job_queue.run_repeating(check_completed_lessons, interval=900, first=60)
            logger.info("✅ Job queue налаштовано")
        else:
            logger.warning("⚠️ Job queue недоступна - нагадування вимкнено")

        logger.info("🚀 Бот запущено!")
        print("🚀 Бот запущено і слухає...")
        print("\n📝 Посилання для реєстрації учнів:")
        print(f"   490 грн: https://t.me/InstructorIFBot?start=register490")
        print(f"   590 грн: https://t.me/InstructorIFBot?start=register590")
        
        import threading
        from http.server import HTTPServer, BaseHTTPRequestHandler
        
        class HealthCheckHandler(BaseHTTPRequestHandler):
            def do_GET(self):
                self.send_response(200)
                self.send_header('Content-type', 'text/plain')
                self.end_headers()
                self.wfile.write(b'Bot is running!')
            
            def log_message(self, format, *args):
                pass
        
        port = int(os.environ.get('PORT', 8080))
        server = HTTPServer(('0.0.0.0', port), HealthCheckHandler)
        server_thread = threading.Thread(target=server.serve_forever, daemon=True)
        server_thread.start()
        logger.info(f"🌐 HTTP сервер запущено на порту {port}")
        print(f"🌐 HTTP сервер запущено на порту {port}")
        
        app.run_polling(drop_pending_updates=True, stop_signals=None)
    
    except Exception as e:
        logger.error(f"Critical error: {e}", exc_info=True)
        raise

if __name__ == "__main__":
    main()

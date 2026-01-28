# import_excel_handler.py
import logging
from io import BytesIO
from typing import List, Dict, Any

from openpyxl import load_workbook
from telegram import Update
from telegram.ext import ContextTypes

logger = logging.getLogger(__name__)

def _normalize_header(h: str) -> str:
    if not h:
        return ""
    s = str(h).strip().lower()
    mapping = {
        "id": "id",
        "дата": "date",
        "час": "time",
        "інструктор": "instructor_name",
        "інструктор id": "instructor_id",
        "учень": "student_name",
        "ім'я": "name",
        "телефон": "student_phone",
        "telegram id": "student_telegram_id",
        "телеграм id": "student_telegram_id",
        "тариф": "student_tariff",
        "тривалість": "duration",
        "вартість": "earnings",
        "статус": "status",
        "оцінка учня": "rating",
        "коментар": "feedback",
        "дата реєстрації": "created_at",
        "created_at": "created_at",
        "tariff": "student_tariff",
        "price_per_hour": "price_per_hour",
        "transmission_type": "transmission_type",
    }
    return mapping.get(s, s.replace(" ", "_"))

def _sheet_to_dicts(ws) -> List[Dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [ _normalize_header(c) if c is not None else "" for c in rows[0] ]
    res = []
    for row in rows[1:]:
        if all([cell is None for cell in row]):
            continue
        d = {}
        for i, cell in enumerate(row):
            key = headers[i] if i < len(headers) else f"col_{i}"
            # try convert numeric IDs to int
            if key in ('id', 'instructor_id', 'student_telegram_id') and cell is not None:
                try:
                    d[key] = int(cell)
                except:
                    d[key] = cell
            else:
                d[key] = cell
        res.append(d)
    return res

async def import_from_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    from bot import is_admin  # Local import to avoid circular dependency
    user_id = update.message.from_user.id
    if not is_admin(user_id):
        await update.message.reply_text("❌ У вас немає доступу до цієї функції.")
        return

    mode = "merge"
    if context.args and str(context.args[0]).lower() == "clear":
        mode = "clear"

    doc = update.message.document
    if not doc:
        await update.message.reply_text(
            "📂 Надішліть файл Excel (.xlsx) у відповідь на команду або прикріпіть файл до повідомлення.\n"
            "Приклад: /import_excel clear  (щоб повністю очистити БД, потім відновити)"
        )
        return

    if not doc.file_name.lower().endswith((".xlsx", ".xlsm", ".xltx")):
        await update.message.reply_text("❌ Потрібен файл у форматі .xlsx (Excel).")
        return

    await update.message.reply_text("⏳ Завантажую та парсю Excel. Зачекайте...")

    try:
        file = await doc.get_file()
        bio = BytesIO()
        await file.download(out=bio)
        bio.seek(0)
        wb = load_workbook(bio, data_only=True)

        instructors = []
        students = []
        lessons = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            name_lower = sheet_name.strip().lower()
            if "інструктор" in name_lower or "instructor" in name_lower:
                instructors = _sheet_to_dicts(ws)
            elif "учн" in name_lower or "student" in name_lower or "учні" in name_lower:
                students = _sheet_to_dicts(ws)
            elif "урок" in name_lower or "lesson" in name_lower:
                lessons = _sheet_to_dicts(ws)
            else:
                hdrs = [ (c.value or "").lower() if c.value else "" for c in ws[1] ]
                hdrs_join = " ".join(hdrs)
                if "дата" in hdrs_join and ("інструктор" in hdrs_join or "учень" in hdrs_join):
                    lessons = _sheet_to_dicts(ws)
                elif "ім'я" in hdrs_join and ("тариф" in hdrs_join or "telegram id" in hdrs_join):
                    students = _sheet_to_dicts(ws)

        from database_imports import import_instructors, import_students, import_lessons
        
        ok_ins = True
        if instructors:
            ok_ins = import_instructors(instructors, clear=(mode == "clear"))
        elif mode == "clear":
            from database import get_db
            with get_db() as conn:
                cur = conn.cursor()
                cur.execute("DELETE FROM lessons")
                cur.execute("DELETE FROM students")
                cur.execute("DELETE FROM instructors")
                conn.commit()
            ok_ins = True

        ok_students = import_students(students) if students else True
        ok_lessons = import_lessons(lessons) if lessons else True

        report = []
        report.append(f"✅ Інструктори: {len(instructors)} - {'OK' if ok_ins else 'FAILED'}")
        report.append(f"✅ Учні: {len(students)} - {'OK' if ok_students else 'FAILED'}")
        report.append(f"✅ Уроки: {len(lessons)} - {'OK' if ok_lessons else 'FAILED'}")

        await update.message.reply_text("\n".join(report))

    except Exception as e:
        logger.error(f"Error in import_from_excel: {e}", exc_info=True)
        await update.message.reply_text(f"❌ Помилка при імпорті: {e}")
